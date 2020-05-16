from PyQt5.QtWidgets import QApplication, QMainWindow, QAction, QLineEdit, QPushButton, QInputDialog, QTableWidget \
    , QHeaderView, QLabel, QCompleter, QComboBox, QRadioButton, QMessageBox, QTextEdit, QDialog, QFileDialog, \
    QTableWidgetItem \
    , QAbstractItemView, QMenu,QCheckBox
from PyQt5.QtCore import QSize, Qt, QSortFilterProxyModel, QTimer
from PyQt5.QtGui import QIcon
import sys
from base_method import BasePage
from appium import webdriver
import os
import pickle
import time
import inspect
from configparser import ConfigParser
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
from testHtmlReport import TestHtmlReport

def send_test_report(email_service, username, password, receves, filename):
    sender = username
    receivers = receves  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱
    # 创建一个带附件的实例
    message = MIMEMultipart()
    message['From'] = username
    message['To'] = receves[0]
    subject = '自动化测试报告'
    message['Subject'] = Header(subject, 'utf-8')

    # 邮件正文内容
    message.attach(MIMEText('详情请参见附件:', 'plain', 'utf-8'))

    # 构造附件1，传送当前目录下的 test.txt 文件
    att1 = MIMEText(open(filename, 'rb').read(), 'base64', 'utf-8')
    att1["Content-Type"] = 'application/octet-stream'
    # 这里的filename可以任意写，写什么名字，邮件中显示什么名字
    att1["Content-Disposition"] = 'attachment; filename={}'.format(filename)
    message.attach(att1)
    smtp = smtplib.SMTP()
    smtp.connect(email_service, 25)  # 25 为 SMTP 端口号
    smtp.login(username, password)
    smtp.sendmail(sender, receivers, message.as_string())


class MyQTimer(QTimer):
    def __new__(cls, *args, **kwargs):
        if not hasattr(MyQTimer, "_instance"):
            MyQTimer._instance = super(QTimer, cls).__new__(cls)
        return MyQTimer._instance


class Config:
    def __init__(self):
        self.conf = ConfigParser()
        if not os.path.exists(os.path.join(os.path.dirname(__file__), "LocatorsObject.ini")):
            with open(os.path.join(os.path.dirname(__file__), "LocatorsObject.ini"), "w+") as f:
                conf = ConfigParser()
                conf.read(filenames="LocatorsObject.ini")
                conf.add_section("test_one_page")
                conf.set("test_one_page", "username_one", "id>kw")
                conf.set("test_one_page", "passwordo_one", "id>su")
                conf.set("test_one_page", "login_one", "id>submit")
                conf.add_section("test_two_page")
                conf.set("test_two_page", "username_two", "id>kw")
                conf.set("test_two_page", "password_two", "id>su")
                conf.set("test_two_page", "login_two", "id>submit")
                conf.write(f)
        self.conf.read(filenames="LocatorsObject.ini",encoding="utf-8")

    def get_all_sections(self):
        return self.conf.sections()

    def get_all_options(self, section):
        return self.conf.options(section)

    def back_locator_tuple(self, section, option):
        locator_name = self.conf.get(section, option)
        return tuple(locator_name.split(">"))


class ExtendedComboBox(QComboBox):
    def __init__(self, parent=None):
        super(ExtendedComboBox, self).__init__(parent)

        self.setFocusPolicy(Qt.StrongFocus)
        self.setEditable(True)

        # add a filter model to filter matching items
        self.pFilterModel = QSortFilterProxyModel(self)
        self.pFilterModel.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.pFilterModel.setSourceModel(self.model())

        # add a completer, which uses the filter model
        self.completer = QCompleter(self.pFilterModel, self)
        # always show all (filtered) completions
        self.completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
        self.setCompleter(self.completer)

        # connect signals
        self.lineEdit().textEdited.connect(self.pFilterModel.setFilterFixedString)
        self.completer.activated.connect(self.on_completer_activated)

    # on selection of an item from the completer, select the corresponding item from combobox
    def on_completer_activated(self, text):
        if text:
            index = self.findText(text)
            self.setCurrentIndex(index)
            self.activated[str].emit(self.itemText(index))

    # on model change, update the models of the filter and completer as well
    def setModel(self, model):
        super(ExtendedComboBox, self).setModel(model)
        self.pFilterModel.setSourceModel(model)
        self.completer.setModel(self.pFilterModel)

    # on model column change, update the model column of the filter and completer as well
    def setModelColumn(self, column):
        self.completer.setCompletionColumn(column)
        self.pFilterModel.setFilterKeyColumn(column)
        super(ExtendedComboBox, self).setModelColumn(column)


class TestCase:
    def __init__(self, title, package,activity, step, page, locator_name, index, data, data_two, data_three, data_transfer, exp,
                 assertmethod):
        self.title = title
        self.package = package
        self.activity = activity

        self.step = step
        self.page = page
        self.locator_name = locator_name

        self.index = index
        self.data = data
        self.data_two = data_two
        self.data_three = data_three
        self.data_transfer = data_transfer

        self.exp = exp

        self.assertmethod = assertmethod


class MainInit(QMainWindow):
    def __init__(self):
        super().__init__()
        self.line_edit_width = int(QApplication.desktop().width() / 4)
        self.table_width = int(QApplication.desktop().width() * 2 / 5)
        self.table_row_height = int(QApplication.desktop().height() * 2 / 3)
        self.global_para = ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                            "", "", "", "", "", "", ""]
        self.line_edit_height = 25
        self.btn_width = 75
        self.btn_height = 25
        self.page_combox_list = []
        self.page_combox_text_list = []
        self.locator_name_combox_list = []
        self.locator_name_combox_text_list = []
        self.step_combox_list = []
        self.step_combox_text_list = []
        self.data_name_list = []
        self.data_name_text_list = []
        self.data_value_list = []
        self.data_value_text_list = []
        self.data_value_two_list = []
        self.data_value_two_text_list = []
        self.data_value_three_list = []
        self.data_value_three_text_list = []
        self.data_output_value_list = []
        self.data_transfer_list = []
        self.data_transfer_text_list = []
        self.result_success = "用例执行通过"
        self.result_success_num = 0
        self.result_fail_num = 0
        self.result_error_num = 0
        self.email_status = False
        self.excute_all_status = 1
        self.driver_name = ""
        self.result_fail = "用例执行失败"
        self.result_error = "用例执行异常"
        self.default_teardown_value = "每个用例执行完关闭app"
        self.default_setup_value = "添加登录操作"
        self.initUI()
        self.setWindowTitle("APP自动化测试工具")
        self.setWindowIcon(QIcon("a.ico"))
        self.setFixedSize(QApplication.desktop().width(), QApplication.desktop().height())
        self.showMaximized()

    def read_config(self):
        self.cf = Config()
        self.sections_list = self.cf.get_all_sections()
        self.options_dict = {}
        for section in self.sections_list:
            self.options_dict[section] = self.cf.get_all_options(section)
        self.sections_list.append("NONE")
        self.options_dict["NONE"] = "NULL"

    def initUI(self):

        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu("文件")
        test_service_url_menu = menu_bar.addMenu("全局配置")
        locate_value_word_menu = menu_bar.addMenu("批量导入导出")
        use_information_menu = menu_bar.addMenu("帮助文档")

        open_test_case_action = QAction("打开测试用例", self)
        open_test_case_action.triggered.connect(self.open_test_case_action_method)
        save_test_case_action = QAction("保存测试用例", self)
        save_test_case_action.triggered.connect(self.save_test_case_action_method)
        new_test_case_action = QAction("新建测试用例", self)
        new_test_case_action.triggered.connect(self.new_test_case_action_method)
        email_action = QAction("邮箱配置", self)
        email_action.triggered.connect(self.email_action_method)
        test_service_url_action = QAction("设备名称和手机版本配置", self)
        test_service_url_action.triggered.connect(self.test_service_url_method)
        mysql_data_action = QAction("数据库配置", self)
        mysql_data_action.triggered.connect(self.mysql_data_method)
        login_information_action = QAction("登录账号信息配置", self)
        login_information_action.triggered.connect(self.login_information_method)
        locate_value_word_action = QAction("查看定位元素文档", self)
        locate_value_word_action.triggered.connect(self.locate_value_word_method)
        user_information_action = QAction("按钮使用说明", self)
        user_information_action.triggered.connect(self.user_information_method)
        import_action = QAction("批量导出测试文件到EXCEL", self)
        import_action.triggered.connect(self.import_method)
        unimport_action = QAction("批量更新测试数据文件", self)
        unimport_action.triggered.connect(self.unimport_method)

        file_menu.addAction(open_test_case_action)
        file_menu.addAction(save_test_case_action)
        file_menu.addAction(new_test_case_action)
        test_service_url_menu.addAction(test_service_url_action)
        test_service_url_menu.addAction(email_action)
        test_service_url_menu.addAction(mysql_data_action)
        test_service_url_menu.addAction(login_information_action)
        use_information_menu.addAction(user_information_action)
        use_information_menu.addAction(locate_value_word_action)
        locate_value_word_menu.addAction(import_action)
        locate_value_word_menu.addAction(unimport_action)

        self.tool = self.addToolBar("执行单个测试用例")

        self.single_excute_action = QAction("执行单个测试用例", self)
        self.single_excute_action.triggered.connect(self.single_excute_action_method)
        self.all_excute_action = QAction("执行所有测试用例", self)
        self.all_excute_action.triggered.connect(self.all_excute_action_method)
        self.setup_action = QAction("添加前置条件", self)
        self.setup_action.triggered.connect(self.setup_action_method)
        self.teardown_action = QAction("添加后置处理", self)
        self.teardown_action.triggered.connect(self.teardown_action_method)
        self.view_result_action = QAction("查看执行结果", self)
        self.view_result_action.triggered.connect(self.view_result_action_method)
        self.view_log_action = QAction("查看执行日志", self)
        self.view_log_action.triggered.connect(self.view_log_action_method)

        self.statu = self.statusBar()
        self.statu.showMessage("ready")

        self.tool.addAction(self.single_excute_action)
        self.tool.addAction(self.all_excute_action)
        self.tool.addAction(self.setup_action)
        self.tool.addAction(self.teardown_action)
        self.tool.addAction(self.view_result_action)
        self.tool.addAction(self.view_log_action)

        self.title_line_edit = QLineEdit(self)
        self.title_line_edit.setPlaceholderText("请输入测试标题")
        self.title_line_edit.adjustSize()
        self.title_line_edit.resize(self.line_edit_width, self.line_edit_height)

        self.package_line_edit = QLineEdit(self)
        self.package_line_edit.setPlaceholderText("请输入APP包名")
        self.package_line_edit.resize(int(self.line_edit_width/2), self.line_edit_height)

        self.activity_line_edit = QLineEdit(self)
        self.activity_line_edit.setPlaceholderText("请输入APP入口名称")
        self.activity_line_edit.resize(int(self.line_edit_width/2), self.line_edit_height)



        self.fail_rerun_combox = QComboBox(self)
        self.fail_rerun_combox.resize(75, 25)
        self.fail_rerun_combox.addItems(["0", "1", "2", "3", "4"])
        self.fail_rerun_combox.setToolTip("失败重跑次数")

        self.add_step_btn = QPushButton("添加步骤", self)
        self.add_step_btn.resize(self.btn_width, self.btn_height)
        self.add_step_btn.clicked.connect(self.add_step_btn_method)

        self.sub_step_btn = QPushButton("删除步骤", self)
        self.sub_step_btn.resize(self.btn_width, self.btn_height)
        self.sub_step_btn.clicked.connect(self.sub_step_btn_method)

        self.data_table = QTableWidget(self)
        self.table_row = 1
        self.data_table.setRowCount(self.table_row)
        self.data_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.data_table.setColumnCount(6)
        self.data_table.setHorizontalHeaderLabels(["步骤名称", "参数一", "参数二", "参数三", "输出数据", "数据传递"])
        # self.data_table.resizeColumnsToContents()
        # self.data_table.resizeRowsToContents()
        self.data_name_list.append(QLineEdit("1"))
        self.data_name_list[0].setReadOnly(True)
        self.data_value_list.append(self.create_Qlineedit_object())
        self.data_value_list[0].editingFinished.connect(self.data_value_list_method)
        data_value_two_edit = QLineEdit()
        self.data_value_two_list.append(self.create_Qlineedit_object())
        self.data_value_two_list[0].editingFinished.connect(self.data_value_two_list_method)
        self.data_value_two_list[0].setReadOnly(True)
        self.data_value_three_list.append(self.create_Qlineedit_object())
        self.data_value_three_list[0].setReadOnly(True)
        self.data_output_value_list.append(self.create_Qlineedit_object())
        self.data_output_value_list[0].setReadOnly(True)
        self.data_output_value_list[0].setContextMenuPolicy(Qt.CustomContextMenu)
        self.data_output_value_list[0].customContextMenuRequested.connect(self.inset_and_delete_action)
        self.data_transfer_list.append(QLineEdit())
        # self.data_value_list[0].setReadOnly(True)
        self.data_value_list[0].setAlignment(Qt.AlignCenter)
        self.data_value_two_list[0].setAlignment(Qt.AlignCenter)
        self.data_value_three_list[0].setAlignment(Qt.AlignCenter)
        self.data_name_list[0].setAlignment(Qt.AlignCenter)
        self.data_output_value_list[0].setAlignment(Qt.AlignCenter)
        self.data_transfer_list[0].setAlignment(Qt.AlignCenter)
        self.data_table.setCellWidget(0, 0, self.data_name_list[0])
        self.data_table.setCellWidget(0, 1, self.data_value_list[0])
        self.data_table.setCellWidget(0, 2, self.data_value_two_list[0])
        self.data_table.setCellWidget(0, 3, self.data_value_three_list[0])
        self.data_table.setCellWidget(0, 4, self.data_output_value_list[0])
        self.data_table.setCellWidget(0, 5, self.data_transfer_list[0])

        self.data_table.resize(self.table_width, self.table_row_height)
        self.data_table.verticalHeader().setVisible(False)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.data_table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.data_table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.exp_line_edit = QLineEdit(self)
        self.exp_line_edit.setPlaceholderText("预期结果")
        self.exp_line_edit.resize(self.line_edit_width, self.line_edit_height)

        self.assert_method_combox = QComboBox(self)
        self.assert_method_combox.resize(self.line_edit_width, self.line_edit_height)
        self.assert_method_combox.addItems(["相等", "不相等"])

        self.act_line_edit = QLineEdit(self)
        self.act_line_edit.setPlaceholderText("实际结果")
        self.act_line_edit.resize(self.line_edit_width, self.line_edit_height)
        self.act_line_edit.setReadOnly(True)

        self.result_label = QLabel(self)
        self.result_label.resize(100, 100)
        self.result_label.setAlignment(Qt.AlignCenter)
        self.result_label.setVisible(False)
        self.read_config()

        self.steps_table = QTableWidget(self)
        self.steps_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.steps_table.verticalHeader().setVisible(False)
        self.steps_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.steps_table_row = 1
        self.steps_table_column = 3
        self.steps_table.setRowCount(self.table_row)
        self.steps_table.setColumnCount(self.steps_table_column)
        self.horizontal_title_list = ["操作步骤", "所属界面", "界面元素名称"]
        self.steps_table.resize(self.table_width, self.table_row_height)
        self.steps_table.setHorizontalHeaderLabels(self.horizontal_title_list)
        self.page_combox_list.append(self.create_Qcombox_object())

        self.page_combox_list[0].addItems(self.sections_list)
        self.page_combox_list[0].currentIndexChanged.connect(self.page_combox_list_method)
        self.locator_name_combox_list.append(self.create_Qcombox_object())
        self.locator_name_combox_list[0].addItems(self.options_dict.get(self.sections_list[0]))
        self.step_combox_list.append(self.create_Qcombox_object())
        self.driver_false = BasePage(driver="driver", title="test.log")
        self.step_combox_list[0].addItems(list(self.driver_false.back_method_dict().keys()))
        self.steps_table.setCellWidget(0, 0, self.step_combox_list[0])
        self.steps_table.setCellWidget(0, 1, self.page_combox_list[0])
        self.steps_table.setCellWidget(0, 2, self.locator_name_combox_list[0])
        self.chrome_radio = QRadioButton("android", self)
        self.chrome_radio.setChecked(True)
        self.chrome_radio.resize(60, 25)
        self.ie_radio = QCheckBox("H5", self)
        self.ie_radio.resize(60, 25)
        self.firefox_radio = QRadioButton("ios", self)
        self.firefox_radio.resize(60, 25)
        self.day_radio = QPushButton("每天生效", self)
        self.week_radio = QPushButton("每周生效", self)
        self.none_radio = QPushButton("全都失效", self)
        self.day_radio.resize(70, 25)
        self.week_radio.resize(70, 25)
        self.none_radio.resize(70, 25)
        self.none_radio.setChecked(True)
        self.day_radio.pressed.connect(self.day_radio_method)
        self.week_radio.pressed.connect(self.week_radio_method)
        self.none_radio.pressed.connect(self.none_radio_method)
        self.hour_combox = QComboBox(self)
        self.hour_combox.addItems(
            ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19",
             "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30",
             "31", "32", "33", "34", "35", "36", "37", "38", "39", "40", "41", "42", "43", "44", "45", "46", "47", "48",
             "49", "50", "51", "52", "53", "54", "55", "56", "57", "58", "59", "0"])
        self.day_combox = QComboBox(self)
        self.day_combox.resize(70, 25)
        self.week_combox = QComboBox(self)
        self.week_combox.resize(70, 25)
        self.hour_combox.resize(70, 25)
        self.day_combox.addItems(
            ["1点", "2点", "3点", "4点", "5点", "6点", "7点", "8点", "9点", "10点", "11点", "12点", "13点", "14点", "15点", "16点",
             "17点", "18点", "19点",
             "20点", "21点", "22点", "23点", "0点"])
        self.week_combox.addItems(["周1", "周2", "周3", "周4", "周5", "周6", "周0"])
        self.timer = MyQTimer(self)
        self.timer.timeout.connect(self.timer_excute_method)
        self.timer.timeout.connect(self.fail_and_error_reexcute)  # 超时执行失败和错误的测试用例
        self.timer.start(1000 * 60)
        self.timer_status = 0

    def fail_and_error_reexcute(self):
        if self.excute_all_status == 0:
            self.all_excute_action_four()

    def create_Qcombox_object(self):
        combox = ExtendedComboBox()
        combox.currentTextChanged.connect(self.hover_combox_display_method)
        return combox

    def hover_combox_display_method(self):
        self.sender().setToolTip(self.sender().currentText())

    def create_Qlineedit_object(self):
        edit = QLineEdit()
        edit.textChanged.connect(self.hover_display_method)
        return edit

    def hover_display_method(self):
        self.sender().setToolTip(self.sender().text())

    def import_method(self):
        base_dir = os.path.join(os.path.dirname(__file__), "test_case_object")
        path = QFileDialog.getExistingDirectory(self, "请选择执行目录", base_dir)
        if path:
            if os.path.exists(os.path.join(path, os.path.basename(path) + "模块测试用例集.xlsx")):
                try:
                    os.remove(os.path.join(path, os.path.basename(path) + "模块测试用例集.xlsx"))
                except Exception:
                    QMessageBox.information(self, "提示", "请先关闭该文件在执行导出操作", QMessageBox.Yes)
                    return None
            wb_create = Workbook()
            sheet_1 = wb_create.active
            sheet_1.cell(1, 1).value = "用例标题"
            sheet_1.cell(1, 2).value = "app包名"
            sheet_1.cell(1, 3).value = "app入口名"
            sheet_1.cell(1, 4).value = "步骤"
            sheet_1.cell(1, 5).value = "预期结果"
            sheet_1.cell(1, 6).value = "断言方式"
            sheet_1.cell(1, 7).value = "文件路径"
            sum = 1
            for object_path, dirs, files in os.walk(path):
                object_path = object_path
                for file in files:
                    if file.endswith(".app"):
                        sum = sum + 1
                        excute_path = os.path.join(object_path, file)
                        test_case = pickle.load(open(excute_path, "rb"))

                        sheet_1.cell(sum, 1).value = test_case.title
                        sheet_1.cell(sum, 2).value = test_case.package
                        sheet_1.cell(sum, 3).value = test_case.activity
                        step_list = []
                        for i in range(0, len(test_case.index)):
                            step_list.append([test_case.index[i], test_case.step[i], test_case.page[i],
                                              test_case.locator_name[i], test_case.data[i], test_case.data_two[i],
                                              test_case.data_three[i], test_case.data_transfer[i]])
                        sheet_1.cell(sum, 4).value = str(step_list)
                        sheet_1.cell(sum, 5).value = test_case.exp
                        sheet_1.cell(sum, 6).value = test_case.assertmethod
                        sheet_1.cell(sum, 7).value = excute_path
            wb_create.save(os.path.join(path, os.path.basename(path) + "模块测试用例集.xlsx"))
            QMessageBox.information(self, "提示", "批量导出EXCEL文件成功", QMessageBox.Yes)

    def unimport_method(self):
        path = QFileDialog.getOpenFileName(self, "打开测试用例表格",
                                               os.path.join(os.path.dirname(__file__), "test_case_object"), "*.xlsx")
        if path[0]:
            wb = load_workbook(path[0])
            sheet_1 = wb.active
            max_row = sheet_1.max_row
            index = 2
            try:
                for i in range(0, max_row - 1):
                    file_path = sheet_1.cell(index, 7).value
                    base_key_list = []
                    page_list = []
                    locator_name_list = []
                    index_list = []
                    data_one_list = []
                    data_two_list = []
                    data_three_list = []
                    data_transfer_list = []
                    try:
                        raw_data = eval(sheet_1.cell(index, 4).value)
                        print(raw_data)
                    except Exception:
                        QMessageBox.information(self, "提示", "第" + str(index) + "行第4列的数据必须是列表", QMessageBox.Yes)
                        return None
                    for j in range(0, len(raw_data)):
                        try:
                            base_key_list.append(raw_data[j][1])
                            page_list.append(raw_data[j][2])
                            locator_name_list.append(raw_data[j][3])
                            index_list.append(raw_data[j][0])
                            data_one_list.append(raw_data[j][4])
                            data_two_list.append(raw_data[j][5])
                            data_three_list.append(raw_data[j][6])
                            data_transfer_list.append(raw_data[j][7])
                        except Exception:
                            QMessageBox.information(self, "提示", "第" + str(index) + "行第4列中第" + str(j + 1) + "个的数据必须是列表",
                                                    QMessageBox.Yes)
                            return None
                    test_case = TestCase(sheet_1.cell(index, 1).value,sheet_1.cell(index, 2).value,sheet_1.cell(index,3).value,base_key_list,
                                         page_list, locator_name_list,
                                         index_list,
                                         data_one_list, data_two_list,
                                         data_three_list,
                                         data_transfer_list, sheet_1.cell(index, 5).value,
                                         sheet_1.cell(index, 6).value)
                    pickle.dump(test_case, open(file_path, "wb"))
                    index = index + 1

            except Exception:
                QMessageBox.information(self, "提示", "测试用例{}：条| 获取文件最大行为{}：行".format(len(raw_data), max_row),
                                        QMessageBox.Yes)
                return None

            QMessageBox.information(self, "提示", "更新测试数据文件成功", QMessageBox.Yes)

    def timer_excute_method(self):  # 执行定时任务方法
        local_m = int(time.strftime("%M"))
        local_h = int(time.strftime("%H"))
        local_d = int(time.strftime("%w"))
        if not self.hour_combox.isEnabled():
            if not self.week_combox.isVisible():
                if local_m == self.select_m and local_h == self.select_h:
                    self.all_excute_action_three(timing_dir)
            else:
                if local_m == self.select_m and local_h == self.select_h and local_d == self.select_d:
                    self.all_excute_action_three(timing_dir)

    def day_radio_method(self):
        reply = QMessageBox.question(self, "提示", "你确定要生效这个定时任务吗？", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            base_dir = os.path.join(os.path.dirname(__file__), "test_case_object")
            path = QFileDialog.getExistingDirectory(self, "请选择执行目录", base_dir)
            if path:
                global timing_dir
                self.select_m = int(self.hour_combox.currentText())
                self.select_h = int(self.day_combox.currentText()[0:-1])
                self.day_combox.setDisabled(True)
                self.day_radio.setDisabled(True)
                self.week_radio.setDisabled(True)
                self.hour_combox.setDisabled(True)
                self.week_combox.setVisible(False)
                timing_dir = path
                QMessageBox.information(self, "提示", "生效成功", QMessageBox.Yes)
            else:
                QMessageBox.information(self, "提示", "没有选择执行目录生效失败", QMessageBox.Yes)

    def week_radio_method(self):
        reply = QMessageBox.question(self, "提示", "你确定要生效这个定时任务吗？", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            base_dir = os.path.join(os.path.dirname(__file__), "test_case_object")
            path = QFileDialog.getExistingDirectory(self, "请选择执行目录", base_dir)
            if path:
                global timing_dir
                self.select_m = int(self.hour_combox.currentText())
                self.select_h = int(self.day_combox.currentText()[0:-1])
                self.select_d = int(self.week_combox.currentText()[1:])
                self.day_radio.setDisabled(True)
                self.week_radio.setDisabled(True)
                self.hour_combox.setDisabled(True)
                self.day_combox.setDisabled(True)
                self.week_combox.setDisabled(True)
                self.week_combox.setVisible(True)
                timing_dir = path
                QMessageBox.information(self, "提示", "生效成功", QMessageBox.Yes)
            else:
                QMessageBox.information(self, "提示", "没有选择执行目录生效失败", QMessageBox.Yes)

    def none_radio_method(self):
        if not self.hour_combox.isEnabled():
            reply = QMessageBox.question(self, "提示", "你确定要取消这个定时任务吗？", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.timer_status = 0
                # self.timer.start(1000*60)
                self.day_combox.setDisabled(False)
                self.day_radio.setDisabled(False)
                self.week_radio.setDisabled(False)
                self.week_combox.setDisabled(False)
                self.hour_combox.setDisabled(False)
                self.week_combox.setVisible(True)
        else:
            QMessageBox.information(self, "提示", "没有定时任务被执行", QMessageBox.Yes)

    def all_excute_action_four(self):  # 用例失败重跑几次
        rerun_num = int(self.fail_rerun_combox.currentText())
        self.excute_all_status = 1
        if rerun_num != 0:
            for i in range(0, rerun_num):
                self.result_success_num = 0
                self.result_fail_num = 0
                self.result_error_num = 0
                self.excute_time = time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(time.time()))
                self.start_time = time.time()
                for object_path, dirs, files in os.walk(
                        os.path.join(os.path.dirname(__file__), "error_and_fail_test_case")):
                    object_path = object_path
                    self.basename = os.path.basename(object_path)
                    for file in files:
                        if file.endswith(".app"):
                            excute_path = os.path.join(object_path, file)
                            test_case = pickle.load(open(excute_path, "rb"))
                            self.title_line_edit.setText(test_case.title)
                            self.package_line_edit.setText(test_case.package)
                            self.activity_line_edit.setText(test_case.activity)
                            self.exp_line_edit.setText(test_case.exp)
                            self.list_clear()
                            self.steps_table_row = len(test_case.step)
                            self.table_row = len(test_case.step)
                            self.data_table.setRowCount(self.table_row)
                            self.steps_table.setRowCount(self.steps_table_row)

                            for i in range(0, len(test_case.step)):
                                self.data_name_list.append(QLineEdit(str(i + 1)))

                                self.data_name_list[i].setReadOnly(True)

                                self.data_name_list[i].setAlignment(Qt.AlignCenter)

                                self.data_value_list.append(QLineEdit())

                                self.data_value_list[i].setAlignment(Qt.AlignCenter)
                                self.data_value_two_list.append(QLineEdit())

                                self.data_value_two_list[i].setAlignment(Qt.AlignCenter)
                                self.data_value_three_list.append(QLineEdit())
                                self.data_value_three_list[i].setAlignment(Qt.AlignCenter)
                                self.data_output_value_list.append(QLineEdit())
                                self.data_output_value_list[i].setReadOnly(True)
                                self.data_output_value_list[i].setAlignment(Qt.AlignCenter)
                                self.data_transfer_list.append(QLineEdit())

                                self.data_transfer_list[i].setAlignment(Qt.AlignCenter)
                                self.data_transfer_list[i].setText(test_case.data_transfer[i])
                                self.data_value_list[i].setText(test_case.data[i])
                                self.data_value_two_list[i].setText(test_case.data_two[i])

                                self.data_value_three_list[i].setText(test_case.data_three[i])
                                if not self.data_value_list[i].text():
                                    self.data_value_two_list[i].setReadOnly(True)
                                if not self.data_value_two_list[i].text():
                                    self.data_value_three_list[i].setReadOnly(True)
                                self.data_value_list[i].editingFinished.connect(self.data_value_list_method)
                                self.data_value_two_list[i].editingFinished.connect(self.data_value_two_list_method)
                                self.data_table.setCellWidget(i, 0, self.data_name_list[i])
                                self.data_table.setCellWidget(i, 1, self.data_value_list[i])

                                self.data_table.setCellWidget(i, 2, self.data_value_two_list[i])
                                self.data_table.setCellWidget(i, 3, self.data_value_three_list[i])
                                self.data_table.setCellWidget(i, 4, self.data_output_value_list[i])
                                self.data_table.setCellWidget(i, 5, self.data_transfer_list[i])
                                self.step_combox_list.append(ExtendedComboBox())
                                self.step_combox_list[i].addItems(list(self.driver_false.back_method_dict().keys()))
                                self.step_combox_list[i].setCurrentText(test_case.step[i])

                                self.page_combox_list.append(ExtendedComboBox())
                                self.page_combox_list[i].addItems(self.sections_list)
                                self.page_combox_list[i].setCurrentText(test_case.page[i])
                                self.page_combox_list[i].currentIndexChanged.connect(self.page_combox_list_method)
                                self.locator_name_combox_list.append(ExtendedComboBox())
                                if self.page_combox_list[i].currentText() != "NONE":
                                    self.locator_name_combox_list[i].addItems(
                                        self.options_dict[self.page_combox_list[i].currentText()])
                                    self.locator_name_combox_list[i].setCurrentText(test_case.locator_name[i])
                                else:
                                    self.locator_name_combox_list[i].addItem(
                                        self.options_dict[self.page_combox_list[i].currentText()])

                                self.steps_table.setCellWidget(i, 0, self.step_combox_list[i])
                                self.steps_table.setCellWidget(i, 1, self.page_combox_list[i])
                                self.steps_table.setCellWidget(i, 2, self.locator_name_combox_list[i])

                            self.single_excute_action_method_three()
                self.end_time = time.time()
                try:
                    with open(
                            os.path.join(os.path.dirname(__file__), "test_case_report",
                                         self.excute_time + "_report.html"),
                            "r") as f:
                        text = f.read()
                    with open(
                            os.path.join(os.path.dirname(__file__), "test_case_report",
                                         self.excute_time + "_report.html"),
                            "w", encoding="utf-8") as f1:
                        all_num = self.result_success_num + self.result_fail_num + self.result_error_num
                        persent_pass = "%.2f%%" % (self.result_success_num / all_num * 100)
                        with open(os.path.join(os.path.dirname(__file__), "SERVICEIP.ini"), "r") as f:
                            ip_name = f.read()
                        f1.write(TestHtmlReport.HTMLHEAD.format(TestHtmlReport.HTMLSCRIPt, self.driver_name, ip_name,
                                                                self.excute_time, self.end_time - self.start_time,
                                                                "共{},通过{},失败{},错误{},通过率={}".
                                                                format(all_num, self.result_success_num,
                                                                       self.result_fail_num, self.result_error_num,
                                                                       persent_pass), self.result_success_num,
                                                                self.result_fail_num, self.result_error_num))
                    with open(
                            os.path.join(os.path.dirname(__file__), "test_case_report",
                                         self.excute_time + "_report.html"),
                            "a+", encoding="utf-8") as f:
                        f.write(text)
                        f.write(TestHtmlReport.HTMLEND)
                    if self.email_status:
                        conf = ConfigParser()
                        conf.read(filenames="EMAIL.ini")
                        receve_list = []
                        file_name = os.path.join(os.path.dirname(__file__), "test_case_report",
                                                 self.excute_time + "_report.html")
                        if conf.get("receve", "receve_user_one"):
                            receve_list.append(conf.get("receve", "receve_user_one"))
                        if conf.get("receve", "receve_user_two"):
                            receve_list.append(conf.get("receve", "receve_user_two"))
                        if conf.get("receve", "receve_user_three"):
                            receve_list.append(conf.get("receve", "receve_user_three"))
                        send_test_report(conf.get("send", "email_service"), conf.get("send", "user_name"),
                                         conf.get("send", "password"), receve_list, file_name)
                except  Exception:
                    pass

    def all_excute_action_three(self, excute_path):  # 定时任务执行方法
        self.result_success_num = 0
        self.result_fail_num = 0
        self.result_error_num = 0
        self.excute_time = time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(time.time()))
        self.start_time = time.time()
        for object_path, dirs, files in os.walk(excute_path):
            object_path = object_path
            self.basename = os.path.basename(object_path)
            for file in files:
                if file.endswith(".app"):
                    excute_path = os.path.join(object_path, file)
                    test_case = pickle.load(open(excute_path, "rb"))
                    self.title_line_edit.setText(test_case.title)
                    self.package_line_edit.setText(test_case.package)
                    self.activity_line_edit.setText(test_case.activity)
                    self.exp_line_edit.setText(test_case.exp)
                    self.list_clear()
                    self.steps_table_row = len(test_case.step)
                    self.table_row = len(test_case.step)
                    self.data_table.setRowCount(self.table_row)
                    self.steps_table.setRowCount(self.steps_table_row)

                    for i in range(0, len(test_case.step)):
                        self.data_name_list.append(QLineEdit(str(i + 1)))

                        self.data_name_list[i].setReadOnly(True)

                        self.data_name_list[i].setAlignment(Qt.AlignCenter)

                        self.data_value_list.append(QLineEdit())

                        self.data_value_list[i].setAlignment(Qt.AlignCenter)
                        self.data_value_two_list.append(QLineEdit())

                        self.data_value_two_list[i].setAlignment(Qt.AlignCenter)
                        self.data_value_three_list.append(QLineEdit())
                        self.data_value_three_list[i].setAlignment(Qt.AlignCenter)
                        self.data_output_value_list.append(QLineEdit())
                        self.data_output_value_list[i].setReadOnly(True)
                        self.data_output_value_list[i].setAlignment(Qt.AlignCenter)
                        self.data_transfer_list.append(QLineEdit())

                        self.data_transfer_list[i].setAlignment(Qt.AlignCenter)
                        self.data_transfer_list[i].setText(test_case.data_transfer[i])
                        self.data_value_list[i].setText(test_case.data[i])
                        self.data_value_two_list[i].setText(test_case.data_two[i])

                        self.data_value_three_list[i].setText(test_case.data_three[i])
                        if not self.data_value_list[i].text():
                            self.data_value_two_list[i].setReadOnly(True)
                        if not self.data_value_two_list[i].text():
                            self.data_value_three_list[i].setReadOnly(True)
                        self.data_value_list[i].editingFinished.connect(self.data_value_list_method)
                        self.data_value_two_list[i].editingFinished.connect(self.data_value_two_list_method)
                        self.data_table.setCellWidget(i, 0, self.data_name_list[i])
                        self.data_table.setCellWidget(i, 1, self.data_value_list[i])

                        self.data_table.setCellWidget(i, 2, self.data_value_two_list[i])
                        self.data_table.setCellWidget(i, 3, self.data_value_three_list[i])
                        self.data_table.setCellWidget(i, 4, self.data_output_value_list[i])
                        self.data_table.setCellWidget(i, 5, self.data_transfer_list[i])
                        self.step_combox_list.append(ExtendedComboBox())
                        self.step_combox_list[i].addItems(list(self.driver_false.back_method_dict().keys()))
                        self.step_combox_list[i].setCurrentText(test_case.step[i])

                        self.page_combox_list.append(ExtendedComboBox())
                        self.page_combox_list[i].addItems(self.sections_list)
                        self.page_combox_list[i].setCurrentText(test_case.page[i])
                        self.page_combox_list[i].currentIndexChanged.connect(self.page_combox_list_method)
                        self.locator_name_combox_list.append(ExtendedComboBox())
                        if self.page_combox_list[i].currentText() != "NONE":
                            self.locator_name_combox_list[i].addItems(
                                self.options_dict[self.page_combox_list[i].currentText()])
                            self.locator_name_combox_list[i].setCurrentText(test_case.locator_name[i])
                        else:
                            self.locator_name_combox_list[i].addItem(
                                self.options_dict[self.page_combox_list[i].currentText()])

                        self.steps_table.setCellWidget(i, 0, self.step_combox_list[i])
                        self.steps_table.setCellWidget(i, 1, self.page_combox_list[i])
                        self.steps_table.setCellWidget(i, 2, self.locator_name_combox_list[i])

                    self.single_excute_action_method_two()
        self.end_time = time.time()
        self.excute_all_status = 0
        try:
            with open(
                    os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                    "r") as f:
                text = f.read()
            with open(
                    os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                    "w", encoding="utf-8") as f1:
                all_num = self.result_success_num + self.result_fail_num + self.result_error_num
                persent_pass = "%.2f%%" % (self.result_success_num / all_num * 100)
                with open(os.path.join(os.path.dirname(__file__), "SERVICEIP.ini"), "r") as f:
                    ip_name = f.read()
                f1.write(TestHtmlReport.HTMLHEAD.format(TestHtmlReport.HTMLSCRIPt, self.driver_name, ip_name,
                                                        self.excute_time, self.end_time - self.start_time,
                                                        "共{},通过{},失败{},错误{},通过率={}".
                                                        format(all_num, self.result_success_num,
                                                               self.result_fail_num, self.result_error_num,
                                                               persent_pass), self.result_success_num,
                                                        self.result_fail_num, self.result_error_num))
            with open(
                    os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                    "a+", encoding="utf-8") as f:
                f.write(text)
                f.write(TestHtmlReport.HTMLEND)
            if self.email_status:
                conf = ConfigParser()
                conf.read(filenames="EMAIL.ini")
                receve_list = []
                file_name = os.path.join(os.path.dirname(__file__), "test_case_report",
                                         self.excute_time + "_report.html")
                if conf.get("receve", "receve_user_one"):
                    receve_list.append(conf.get("receve", "receve_user_one"))
                if conf.get("receve", "receve_user_two"):
                    receve_list.append(conf.get("receve", "receve_user_two"))
                if conf.get("receve", "receve_user_three"):
                    receve_list.append(conf.get("receve", "receve_user_three"))
                send_test_report(conf.get("send", "email_service"), conf.get("send", "user_name"),
                                 conf.get("send", "password"), receve_list, file_name)
        except  Exception:
            pass

    def inset_and_delete_action(self, pos):
        index = self.data_output_value_list.index(self.sender())
        position = index + 1
        menu = QMenu()
        insert_action = menu.addAction("插入一行")
        delete_action = menu.addAction('删除这行')
        # global_one_action = menu.addAction("设为红色")
        # global_two_action = menu.addAction("设为绿色")
        # global_three_action = menu.addAction("设为黄色")
        action = menu.exec_(self.data_output_value_list[index].mapToGlobal(pos))
        if action == insert_action:
            self.steps_table.insertRow(position)
            self.data_table.insertRow(position)
            self.steps_table_row = self.steps_table_row + 1
            # self.steps_table.setRowCount(self.steps_table_row)
            self.table_row = self.table_row + 1
            # self.data_table.setRowCount(self.table_row)
            self.data_name_list.insert(position, QLineEdit(str(position + 1)))
            for i in range(position + 1, len(self.data_name_list)):
                self.data_name_list[i].setText(str(i + 1))
            self.data_name_list[position].setAlignment(Qt.AlignCenter)
            self.data_name_list[position].setReadOnly(True)
            self.data_value_list.insert(position, self.create_Qlineedit_object())
            self.data_value_list[position].setAlignment(Qt.AlignCenter)
            self.data_value_list[position].editingFinished.connect(self.data_value_list_method)
            self.data_value_two_list.insert(position, self.create_Qlineedit_object())
            self.data_value_two_list[position].setAlignment(Qt.AlignCenter)
            self.data_value_two_list[position].setReadOnly(True)
            self.data_value_two_list[position].editingFinished.connect(self.data_value_two_list_method)
            self.data_value_three_list.insert(position, self.create_Qlineedit_object())
            self.data_value_three_list[position].setAlignment(Qt.AlignCenter)
            self.data_value_three_list[position].setReadOnly(True)
            self.data_output_value_list.insert(position, self.create_Qlineedit_object())
            self.data_output_value_list[position].setReadOnly(True)
            self.data_transfer_list.insert(position, QLineEdit())
            self.data_transfer_list[position].setAlignment(Qt.AlignCenter)
            self.data_output_value_list[position].setAlignment(Qt.AlignCenter)
            self.data_output_value_list[position].setContextMenuPolicy(Qt.CustomContextMenu)
            self.data_output_value_list[position].customContextMenuRequested.connect(self.inset_and_delete_action)

            self.step_combox_list.insert(position, self.create_Qcombox_object())
            self.step_combox_list[position].addItems(list(self.driver_false.back_method_dict().keys()))
            self.page_combox_list.insert(position, self.create_Qcombox_object())
            self.page_combox_list[position].addItems(self.sections_list)
            self.page_combox_list[position].currentIndexChanged.connect(self.page_combox_list_method)
            self.locator_name_combox_list.insert(position, self.create_Qcombox_object())
            self.locator_name_combox_list[position].addItems(self.options_dict.get(self.sections_list[0]))

            self.data_table.setCellWidget(position, 0, self.data_name_list[position])
            self.data_table.setCellWidget(position, 1, self.data_value_list[position])
            self.data_table.setCellWidget(position, 2, self.data_value_two_list[position])
            self.data_table.setCellWidget(position, 3, self.data_value_three_list[position])
            self.data_table.setCellWidget(position, 4, self.data_output_value_list[position])
            self.data_table.setCellWidget(position, 5, self.data_transfer_list[position])

            self.steps_table.setCellWidget(position, 0, self.step_combox_list[position])
            self.steps_table.setCellWidget(position, 1, self.page_combox_list[position])
            self.steps_table.setCellWidget(position, 2, self.locator_name_combox_list[position])
        if action == delete_action:
            self.steps_table_row = self.steps_table_row - 1
            self.table_row = self.table_row - 1

            self.step_combox_list.pop(index)
            self.page_combox_list.pop(index)
            self.locator_name_combox_list.pop(index)

            self.data_name_list.pop(index)
            for i in range(index, len(self.data_name_list)):
                self.data_name_list[i].setText(str(i + 1))
            self.data_value_list.pop(index)
            self.data_value_two_list.pop(index)
            self.data_value_three_list.pop(index)
            self.data_output_value_list.pop(index)
            self.data_transfer_list.pop(index)
            self.steps_table.removeRow(index)
            self.data_table.removeRow(index)


    def list_clear(self):
        self.data_name_list.clear()
        self.data_value_list.clear()
        self.data_value_two_list.clear()
        self.data_value_three_list.clear()
        self.data_output_value_list.clear()
        self.data_transfer_list.clear()
        self.step_combox_list.clear()
        self.page_combox_list.clear()
        self.locator_name_combox_list.clear()

    def list_text_clear(self):
        self.data_name_text_list.clear()
        self.data_value_text_list.clear()
        self.data_value_two_text_list.clear()
        self.data_value_three_text_list.clear()
        self.data_transfer_text_list.clear()
        self.step_combox_text_list.clear()
        self.page_combox_text_list.clear()
        self.locator_name_combox_text_list.clear()

    def all_excute_action_method(self):
        self.result_success_num = 0
        self.result_fail_num = 0
        self.result_error_num = 0
        self.excute_time = time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(time.time()))
        self.start_time = time.time()
        base_dir = os.path.join(os.path.dirname(__file__), "test_case_object")
        path = QFileDialog.getExistingDirectory(self, "请选择执行目录", base_dir)
        if path:
            for object_path, dirs, files in os.walk(path):
                object_path = object_path
                self.basename = os.path.basename(object_path)
                for file in files:
                    if file.endswith(".app"):
                        global excute_path
                        excute_path = os.path.join(object_path, file)
                        test_case = pickle.load(open(excute_path, "rb"))
                        self.title_line_edit.setText(test_case.title)
                        self.package_line_edit.setText(test_case.package)
                        self.activity_line_edit.setText(test_case.activity)
                        self.exp_line_edit.setText(test_case.exp)
                        self.list_clear()
                        self.steps_table_row = len(test_case.step)
                        self.table_row = len(test_case.step)
                        self.data_table.setRowCount(self.table_row)
                        self.steps_table.setRowCount(self.steps_table_row)

                        for i in range(0, len(test_case.step)):
                            self.data_name_list.append(QLineEdit(str(i + 1)))

                            self.data_name_list[i].setReadOnly(True)

                            self.data_name_list[i].setAlignment(Qt.AlignCenter)

                            self.data_value_list.append(QLineEdit())

                            self.data_value_list[i].setAlignment(Qt.AlignCenter)
                            self.data_value_two_list.append(QLineEdit())

                            self.data_value_two_list[i].setAlignment(Qt.AlignCenter)
                            self.data_value_three_list.append(QLineEdit())
                            self.data_value_three_list[i].setAlignment(Qt.AlignCenter)
                            self.data_output_value_list.append(QLineEdit())
                            self.data_output_value_list[i].setReadOnly(True)
                            self.data_output_value_list[i].setAlignment(Qt.AlignCenter)
                            self.data_transfer_list.append(QLineEdit())

                            self.data_transfer_list[i].setAlignment(Qt.AlignCenter)
                            self.data_transfer_list[i].setText(test_case.data_transfer[i])
                            self.data_value_list[i].setText(test_case.data[i])
                            self.data_value_two_list[i].setText(test_case.data_two[i])

                            self.data_value_three_list[i].setText(test_case.data_three[i])
                            if not self.data_value_list[i].text():
                                self.data_value_two_list[i].setReadOnly(True)
                            if not self.data_value_two_list[i].text():
                                self.data_value_three_list[i].setReadOnly(True)
                            self.data_value_list[i].editingFinished.connect(self.data_value_list_method)
                            self.data_value_two_list[i].editingFinished.connect(self.data_value_two_list_method)
                            self.data_table.setCellWidget(i, 0, self.data_name_list[i])
                            self.data_table.setCellWidget(i, 1, self.data_value_list[i])

                            self.data_table.setCellWidget(i, 2, self.data_value_two_list[i])
                            self.data_table.setCellWidget(i, 3, self.data_value_three_list[i])
                            self.data_table.setCellWidget(i, 4, self.data_output_value_list[i])
                            self.data_table.setCellWidget(i, 5, self.data_transfer_list[i])
                            self.step_combox_list.append(ExtendedComboBox())
                            self.step_combox_list[i].addItems(list(self.driver_false.back_method_dict().keys()))
                            self.step_combox_list[i].setCurrentText(test_case.step[i])

                            self.page_combox_list.append(ExtendedComboBox())
                            self.page_combox_list[i].addItems(self.sections_list)
                            self.page_combox_list[i].setCurrentText(test_case.page[i])
                            self.page_combox_list[i].currentIndexChanged.connect(self.page_combox_list_method)
                            self.locator_name_combox_list.append(ExtendedComboBox())
                            if self.page_combox_list[i].currentText() != "NONE":
                                self.locator_name_combox_list[i].addItems(
                                    self.options_dict[self.page_combox_list[i].currentText()])
                                self.locator_name_combox_list[i].setCurrentText(test_case.locator_name[i])
                            else:
                                self.locator_name_combox_list[i].addItem(
                                    self.options_dict[self.page_combox_list[i].currentText()])

                            self.steps_table.setCellWidget(i, 0, self.step_combox_list[i])
                            self.steps_table.setCellWidget(i, 1, self.page_combox_list[i])
                            self.steps_table.setCellWidget(i, 2, self.locator_name_combox_list[i])


                        self.single_excute_action_method_two()
            self.end_time = time.time()
            self.excute_all_status = 0
            try:
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "r") as f:
                    text = f.read()
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "w", encoding="utf-8") as f1:
                    all_num = self.result_success_num + self.result_fail_num + self.result_error_num
                    persent_pass = "%.2f%%" % (self.result_success_num / all_num * 100)
                    with open(os.path.join(os.path.dirname(__file__), "SERVICEIP.ini"), "r") as f:
                        ip_name = f.read()
                    f1.write(TestHtmlReport.HTMLHEAD.format(TestHtmlReport.HTMLSCRIPt, self.driver_name, ip_name,
                                                            self.excute_time, self.end_time - self.start_time,
                                                            "共{},通过{},失败{},错误{},通过率={}".
                                                            format(all_num, self.result_success_num,
                                                                   self.result_fail_num, self.result_error_num,
                                                                   persent_pass), self.result_success_num,
                                                            self.result_fail_num, self.result_error_num))
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "a+", encoding="utf-8") as f:
                    f.write(text)
                    f.write(TestHtmlReport.HTMLEND)
                if self.email_status:
                    conf = ConfigParser()
                    conf.read(filenames="EMAIL.ini")
                    receve_list = []
                    file_name = os.path.join(os.path.dirname(__file__), "test_case_report",
                                             self.excute_time + "_report.html")
                    if conf.get("receve", "receve_user_one"):
                        receve_list.append(conf.get("receve", "receve_user_one"))
                    if conf.get("receve", "receve_user_two"):
                        receve_list.append(conf.get("receve", "receve_user_two"))
                    if conf.get("receve", "receve_user_three"):
                        receve_list.append(conf.get("receve", "receve_user_three"))
                    send_test_report(conf.get("send", "email_service"), conf.get("send", "user_name"),
                                     conf.get("send", "password"), receve_list, file_name)

            except  Exception:
                pass

    def add_step_btn_method(self):
        self.steps_table_row = self.steps_table_row + 1
        self.steps_table.setRowCount(self.steps_table_row)
        self.table_row = self.table_row + 1
        self.data_table.setRowCount(self.table_row)
        self.data_name_list.append(QLineEdit(str(self.steps_table_row)))
        self.data_name_list[-1].setAlignment(Qt.AlignCenter)
        self.data_name_list[-1].setReadOnly(True)
        self.data_value_list.append(self.create_Qlineedit_object())
        self.data_value_list[-1].setAlignment(Qt.AlignCenter)
        self.data_value_list[-1].editingFinished.connect(self.data_value_list_method)
        self.data_value_two_list.append(self.create_Qlineedit_object())
        self.data_value_two_list[-1].setAlignment(Qt.AlignCenter)
        self.data_value_two_list[-1].setReadOnly(True)
        self.data_value_two_list[-1].editingFinished.connect(self.data_value_two_list_method)
        self.data_value_three_list.append(self.create_Qlineedit_object())
        self.data_value_three_list[-1].setAlignment(Qt.AlignCenter)
        self.data_value_three_list[-1].setReadOnly(True)
        self.data_output_value_list.append(self.create_Qlineedit_object())
        self.data_output_value_list[-1].setReadOnly(True)
        self.data_transfer_list.append(QLineEdit())
        self.data_transfer_list[-1].setAlignment(Qt.AlignCenter)
        self.data_output_value_list[-1].setAlignment(Qt.AlignCenter)
        self.data_output_value_list[-1].setContextMenuPolicy(Qt.CustomContextMenu)
        self.data_output_value_list[-1].customContextMenuRequested.connect(self.inset_and_delete_action)

        self.step_combox_list.append(self.create_Qcombox_object())
        self.step_combox_list[-1].addItems(list(self.driver_false.back_method_dict().keys()))
        self.page_combox_list.append(self.create_Qcombox_object())
        self.page_combox_list[-1].addItems(self.sections_list)
        self.page_combox_list[-1].currentIndexChanged.connect(self.page_combox_list_method)
        self.locator_name_combox_list.append(self.create_Qcombox_object())
        self.locator_name_combox_list[-1].addItems(self.options_dict[self.sections_list[0]])

        self.data_table.setCellWidget(self.steps_table_row - 1, 0, self.data_name_list[-1])
        self.data_table.setCellWidget(self.steps_table_row - 1, 1, self.data_value_list[-1])
        self.data_table.setCellWidget(self.steps_table_row - 1, 2, self.data_value_two_list[-1])
        self.data_table.setCellWidget(self.steps_table_row - 1, 3, self.data_value_three_list[-1])
        self.data_table.setCellWidget(self.steps_table_row - 1, 4, self.data_output_value_list[-1])
        self.data_table.setCellWidget(self.steps_table_row - 1, 5, self.data_transfer_list[-1])

        self.steps_table.setCellWidget(self.steps_table_row - 1, 0, self.step_combox_list[-1])
        self.steps_table.setCellWidget(self.steps_table_row - 1, 1, self.page_combox_list[-1])

        self.steps_table.setCellWidget(self.steps_table_row - 1, 2, self.locator_name_combox_list[-1])

    def data_value_list_method(self):
        index = self.data_value_list.index(self.sender())
        if self.data_value_list[index].text():
            self.data_value_two_list[index].setReadOnly(False)
        else:
            self.data_value_two_list[index].setReadOnly(True)
            self.data_value_two_list[index].setText("")
            self.data_value_three_list[index].setReadOnly(True)
            self.data_value_three_list[index].setText("")

    def data_value_two_list_method(self):

        index = self.data_value_two_list.index(self.sender())
        if self.data_value_two_list[index].text():
            self.data_value_three_list[index].setReadOnly(False)
        else:
            self.data_value_three_list[index].setReadOnly(True)
            self.data_value_three_list[index].setText("")

    def locate_value_word_method(self):
        if os.path.exists(os.path.join(os.path.dirname(__file__), "LocatorsObject.ini")):
            dialog = QDialog()
            dialog.setWindowTitle("定位元素文档查看")
            locator_edit = QTextEdit(dialog)
            locator_edit.resize(QApplication.desktop().width(), QApplication.desktop().height())
            with open(os.path.join(os.path.dirname(__file__), "LocatorsObject.ini"), "rb+") as f:
                locator_edit.setHtml(f.read())
            locator_edit.setReadOnly(True)
            dialog.showMaximized()
            dialog.exec_()
        else:
            QMessageBox.information(self, "提示", "配置文件不存在", QMessageBox.Ok)

    def page_combox_list_method(self):
        try:
            index = self.page_combox_list.index(self.sender())
            if self.page_combox_list[index].currentText() != "NONE":
                self.locator_name_combox_list[index].clear()
                self.locator_name_combox_list[index].addItems(
                    self.options_dict[self.page_combox_list[index].currentText()])
            else:
                self.locator_name_combox_list[index].clear()
                self.locator_name_combox_list[index].addItem(
                    self.options_dict[self.page_combox_list[index].currentText()])
        except Exception:
            QMessageBox.information(self, "提示", "SECTION不存在", QMessageBox.Ok)

    def login_information_method(self):
        file_path = os.path.join(os.path.dirname(__file__), "LOGIN.ini")
        dialog = QDialog()
        dialog.setWindowTitle("登录账号配置")
        username_edit = QLineEdit(dialog)
        password_edit = QLineEdit(dialog)
        username_edit.setPlaceholderText("请输入登录账号")
        password_edit.setPlaceholderText("请输入登录密码")

        def remove_btn_method():
            dialog.close()

        def submit_btn_method():
            if username_edit.text() and password_edit.text():
                with open(file_path, "w+") as f:
                    f.write(
                        username_edit.text() + ";" + password_edit.text())
                dialog.close()

        submit_btn = QPushButton("确定", dialog)
        submit_btn.clicked.connect(submit_btn_method)
        remove_btn = QPushButton("取消", dialog)
        remove_btn.clicked.connect(remove_btn_method)
        submit_btn.resize(150, 30)
        submit_btn.move(0, 60)
        remove_btn.resize(150, 30)
        remove_btn.move(150, 60)
        username_edit.resize(300, 30)
        password_edit.resize(300, 30)
        password_edit.setEchoMode(QLineEdit.Password)
        password_edit.move(0, 30)
        dialog.resize(300, 90)
        with open(file_path, "r") as f:
            read_string = f.read()
            read_list = read_string.split(";")
            if len(read_list) == 2:
                username_edit.setText(read_list[0])
                password_edit.setText(read_list[1])
        dialog.exec_()

    def mysql_data_method(self):
        file_path = os.path.join(os.path.dirname(__file__), "MYSQL.ini")
        dialog = QDialog()
        dialog.setWindowTitle("数据库配置")
        ip_line_edit = QLineEdit(dialog)
        user_name_edit = QLineEdit(dialog)
        password_edit = QLineEdit(dialog)
        port_edit = QLineEdit(dialog)
        database_edit = QLineEdit(dialog)
        port_edit.setPlaceholderText("请输入端口号")
        database_edit.setPlaceholderText("请输入数据库名称")

        def remove_btn_method():
            dialog.close()

        def submit_btn_method():
            if ip_line_edit.text() and user_name_edit.text() and password_edit.text() and port_edit.text() and database_edit.text():
                with open(file_path, "w+") as f:
                    f.write(
                        ip_line_edit.text() + ";" + user_name_edit.text() + ";" + password_edit.text() + ";" + port_edit.text() + ";" + database_edit.text())
                dialog.close()

        submit_btn = QPushButton("确定", dialog)
        submit_btn.clicked.connect(submit_btn_method)
        remove_btn = QPushButton("取消", dialog)
        remove_btn.clicked.connect(remove_btn_method)
        submit_btn.resize(150, 30)
        submit_btn.move(0, 90)
        remove_btn.resize(150, 30)
        remove_btn.move(150, 90)
        ip_line_edit.resize(300, 30)
        ip_line_edit.setPlaceholderText("请输入测试服务器地址及端口号")
        user_name_edit.resize(300, 30)
        user_name_edit.setPlaceholderText("请输入用户名")
        user_name_edit.move(0, 30)
        password_edit.resize(150, 30)
        password_edit.setPlaceholderText("请输入密码")
        password_edit.setEchoMode(QLineEdit.Password)
        password_edit.move(150, 30)
        port_edit.resize(150, 30)
        database_edit.resize(150, 30)
        port_edit.move(0, 60)
        database_edit.move(150, 60)
        dialog.resize(300, 120)
        with open(file_path, "r") as f:
            read_string = f.read()
            read_list = read_string.split(";")
            if len(read_list) == 5:
                ip_line_edit.setText(read_list[0])
                user_name_edit.setText(read_list[1])
                password_edit.setText(read_list[2])
                port_edit.setText(read_list[3])
                database_edit.setText(read_list[4])
        dialog.exec_()

    def new_test_case_action_method(self):
        self.title_line_edit.setText("")
        self.exp_line_edit.setText("")
        self.table_row = 1
        self.steps_table_row = 1
        self.steps_table.setRowCount(self.table_row)
        self.data_table.setRowCount(self.table_row)
        self.data_value_list[0].setText("")
        self.data_value_two_list[0].setText("")
        self.data_value_two_list[0].setReadOnly(True)
        self.data_value_three_list[0].setReadOnly(True)
        self.data_value_list[0].editingFinished.connect(self.data_value_list_method)
        self.data_value_two_list[0].editingFinished.connect(self.data_value_two_list_method)
        self.data_value_three_list[0].setText("")
        self.data_transfer_list[0].setText("")
        self.data_output_value_list[0].setText("")
        self.list_text_clear()
        data_num = len(self.data_value_list) - 1
        for i in range(0, data_num):
            self.step_combox_list.pop(-1)
            self.page_combox_list.pop(-1)
            self.locator_name_combox_list.pop(-1)
            self.data_name_list.pop(-1)
            self.data_value_list.pop(-1)
            self.data_value_two_list.pop(-1)
            self.data_value_three_list.pop(-1)
            self.data_output_value_list.pop(-1)
            self.data_transfer_list.pop(-1)

    def open_test_case_action_method(self):
        path = QFileDialog.getOpenFileName(self, "打开测试用例",
                                               os.path.join(os.path.dirname(__file__), "test_case_object"), "*.app")
        if path[0]:
            try:
                test_case = pickle.load(open(path[0], "rb"))
                self.title_line_edit.setText(test_case.title)
                self.package_line_edit.setText(test_case.package)
                self.activity_line_edit.setText(test_case.activity)
                self.exp_line_edit.setText(test_case.exp)
                self.list_clear()
                self.steps_table_row = len(test_case.step)
                self.table_row = len(test_case.step)
                self.data_table.setRowCount(self.table_row)
                self.steps_table.setRowCount(self.steps_table_row)

                for i in range(0, len(test_case.step)):
                    self.data_name_list.append(QLineEdit(str(i + 1)))
                    self.data_name_list[i].setReadOnly(True)
                    self.data_name_list[i].setAlignment(Qt.AlignCenter)
                    self.data_value_list.append(self.create_Qlineedit_object())
                    self.data_value_list[i].setAlignment(Qt.AlignCenter)
                    self.data_value_two_list.append(self.create_Qlineedit_object())
                    self.data_value_two_list[i].setAlignment(Qt.AlignCenter)
                    self.data_value_three_list.append(self.create_Qlineedit_object())
                    self.data_value_three_list[i].setAlignment(Qt.AlignCenter)
                    self.data_output_value_list.append(self.create_Qlineedit_object())
                    self.data_output_value_list[i].setReadOnly(True)
                    self.data_output_value_list[i].setAlignment(Qt.AlignCenter)
                    self.data_output_value_list[i].setContextMenuPolicy(Qt.CustomContextMenu)
                    self.data_output_value_list[i].customContextMenuRequested.connect(self.inset_and_delete_action)
                    self.data_transfer_list.append(QLineEdit())
                    self.data_transfer_list[i].setAlignment(Qt.AlignCenter)
                    self.data_transfer_list[i].setText(test_case.data_transfer[i])
                    self.data_value_list[i].setText(test_case.data[i])
                    self.data_value_two_list[i].setText(test_case.data_two[i])
                    self.data_value_three_list[i].setText(test_case.data_three[i])
                    if not self.data_value_list[i].text():
                        self.data_value_two_list[i].setReadOnly(True)
                    if not self.data_value_two_list[i].text():
                        self.data_value_three_list[i].setReadOnly(True)
                    self.data_value_list[i].editingFinished.connect(self.data_value_list_method)
                    self.data_value_two_list[i].editingFinished.connect(self.data_value_two_list_method)
                    self.data_table.setCellWidget(i, 0, self.data_name_list[i])
                    self.data_table.setCellWidget(i, 1, self.data_value_list[i])
                    self.data_table.setCellWidget(i, 2, self.data_value_two_list[i])
                    self.data_table.setCellWidget(i, 3, self.data_value_three_list[i])
                    self.data_table.setCellWidget(i, 4, self.data_output_value_list[i])
                    self.data_table.setCellWidget(i, 5, self.data_transfer_list[i])
                    self.step_combox_list.append(self.create_Qcombox_object())
                    self.step_combox_list[i].addItems(list(self.driver_false.back_method_dict().keys()))
                    self.step_combox_list[i].setCurrentText(test_case.step[i])
                    if test_case.step[i] not in list(self.driver_false.back_method_dict().keys()):
                        QMessageBox.information(self, '提示', "第" + str(i + 1) + "步的关键字不存在", QMessageBox.Yes)
                    self.page_combox_list.append(self.create_Qcombox_object())
                    self.page_combox_list[i].addItems(self.sections_list)
                    self.page_combox_list[i].setCurrentText(test_case.page[i])
                    self.page_combox_list[i].currentIndexChanged.connect(self.page_combox_list_method)
                    self.locator_name_combox_list.append(self.create_Qcombox_object())
                    if self.page_combox_list[i].currentText() != "NONE":
                        self.locator_name_combox_list[i].addItems(
                            self.options_dict.get(self.page_combox_list[i].currentText()))
                        self.locator_name_combox_list[i].setCurrentText(test_case.locator_name[i])
                        if test_case.locator_name[i] not in list(
                                self.options_dict.get(self.page_combox_list[i].currentText())):
                            QMessageBox.information(self, '提示', "第" + str(i + 1) + "步的OPTIONS不存在", QMessageBox.Yes)
                    else:
                        self.locator_name_combox_list[i].addItem(
                            self.options_dict.get(self.page_combox_list[i].currentText()))

                    self.steps_table.setCellWidget(i, 0, self.step_combox_list[i])
                    self.steps_table.setCellWidget(i, 1, self.page_combox_list[i])
                    self.steps_table.setCellWidget(i, 2, self.locator_name_combox_list[i])
                self.list_text_clear()
                self.assert_method_combox.setCurrentText(test_case.assertmethod)
            except Exception as e:
                print("1")
                QMessageBox.information(self, '提示', "第" + str(i + 1) + "步的SECTION不存在", QMessageBox.Yes)
                # infoBox = QMessageBox(self)  ##Message Box that doesn't run
                # infoBox.setIcon(QMessageBox.Information)
                # infoBox.setText("第"+str(i+1)+"步的SECTION不存在")
                # infoBox.setWindowTitle("提示")
                # infoBox.setStandardButtons(QMessageBox.Ok)
                # infoBox.button(QMessageBox.Ok).animateClick(3 * 1000)
                # infoBox.exec_()

    def closeEvent(self, QCloseEvent):
        if not self.hour_combox.isEnabled():
            res = QMessageBox.question(self, '提示', '存在定时任务确定关闭程序？', QMessageBox.Yes | QMessageBox.No)
            if res == QMessageBox.Yes:
                QCloseEvent.accept()
            else:
                QCloseEvent.ignore()
        else:
            QCloseEvent.accept()

    def resizeEvent(self, sizeEvent):
        self.main_width = QApplication.desktop().width()
        self.main_height = QApplication.desktop().height()
        self.title_line_edit.move(int(self.main_width / 60), int(self.main_height / 16))
        self.package_line_edit.move(int(self.main_width / 60 + self.line_edit_width + 10), int(self.main_height / 16))
        self.activity_line_edit.move(int(self.main_width / 60 + self.line_edit_width + 10+self.line_edit_width/2), int(self.main_height / 16))
        self.exp_line_edit.move(int(self.main_width / 60 + self.line_edit_width * 2 + 20), int(self.main_height / 16))
        self.assert_method_combox.move(int(self.main_width / 60), int(self.main_height * 2 / 16))
        self.act_line_edit.move(int(self.main_width / 60 + self.line_edit_width + 10), int(self.main_height * 2 / 16))
        self.chrome_radio.move(int(self.main_width / 60 + self.line_edit_width * 2 + 20),
                               int(self.main_height * 2 / 16))
        self.ie_radio.move(
            int(self.main_width / 60 + self.line_edit_width * 2 + 20 + QApplication.desktop().width() / 8 - 30),
            int(self.main_height * 2 / 16))
        self.firefox_radio.move(
            int(self.main_width / 60 + self.line_edit_width * 2 + 40 + QApplication.desktop().width() / 4 - 60),
            int(self.main_height * 2 / 16))
        self.steps_table.move(int(self.main_width / 60), int(self.main_height * 3 / 16))
        self.data_table.move(int(self.main_width / 60 + self.table_width), int(self.main_height * 3 / 16))
        self.fail_rerun_combox.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height / 16))
        self.add_step_btn.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height * 3 / 16))
        self.sub_step_btn.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height * 3 / 16 + 25))
        self.result_label.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height * 3 / 16 + 50))
        self.none_radio.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height * 3 / 16 + 200))
        self.day_radio.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height * 3 / 16 + 240))
        self.week_radio.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height * 3 / 16 + 280))
        self.hour_combox.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height * 3 / 16 + 320))
        self.day_combox.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height * 3 / 16 + 360))
        self.week_combox.move(int(self.main_width / 60 + self.table_width * 2), int(self.main_height * 3 / 16 + 400))

    def setup_action_method(self):
        path = QFileDialog.getOpenFileName(self, "打开正常登录测试用例",
                                           os.path.join(os.path.dirname(__file__), "test_case_object"), "*.app")
        if path[0]:
            test_case = pickle.load(open(path[0], "rb"))
            self.package_line_edit.setText(test_case.package)
            self.activity_line_edit.setText(test_case.activity)
            self.list_clear()
            self.steps_table_row = len(test_case.step)
            self.table_row = len(test_case.step)
            self.data_table.setRowCount(self.table_row)
            self.steps_table.setRowCount(self.steps_table_row)

            for i in range(0, len(test_case.step)):
                self.data_name_list.append(QLineEdit(str(i + 1)))
                self.data_name_list[i].setReadOnly(True)
                self.data_value_list.append(self.create_Qlineedit_object())
                self.data_value_list[i].setText(test_case.data[i])
                self.data_value_list[i].editingFinished.connect(self.data_value_list_method)
                self.data_value_two_list.append(self.create_Qlineedit_object())
                self.data_value_two_list[i].setText(test_case.data_two[i])
                self.data_value_two_list[i].editingFinished.connect(self.data_value_two_list_method)
                if not self.data_value_list[i].text():
                    self.data_value_two_list[i].setReadOnly(True)
                self.data_value_three_list.append(self.create_Qlineedit_object())
                if not self.data_value_two_list[i].text():
                    self.data_value_three_list[i].setReadOnly(True)
                self.data_value_three_list[i].setText(test_case.data_three[i])
                self.data_output_value_list.append(self.create_Qlineedit_object())
                self.data_output_value_list[i].setAlignment(Qt.AlignCenter)
                self.data_output_value_list[i].setReadOnly(True)
                self.data_output_value_list[i].setContextMenuPolicy(Qt.CustomContextMenu)
                self.data_output_value_list[i].customContextMenuRequested.connect(self.inset_and_delete_action)
                self.data_transfer_list.append(QLineEdit())
                self.data_transfer_list[i].setAlignment(Qt.AlignCenter)
                self.data_transfer_list[i].setText(test_case.data_transfer[i])
                self.data_value_list[i].setAlignment(Qt.AlignCenter)
                self.data_value_two_list[i].setAlignment(Qt.AlignCenter)
                self.data_value_three_list[i].setAlignment(Qt.AlignCenter)
                self.data_name_list[i].setAlignment(Qt.AlignCenter)
                self.data_table.setCellWidget(i, 0, self.data_name_list[i])
                self.data_table.setCellWidget(i, 1, self.data_value_list[i])
                self.data_table.setCellWidget(i, 2, self.data_value_two_list[i])
                self.data_table.setCellWidget(i, 3, self.data_value_three_list[i])
                self.data_table.setCellWidget(i, 4, self.data_output_value_list[i])
                self.data_table.setCellWidget(i, 5, self.data_transfer_list[i])

                self.step_combox_list.append(self.create_Qcombox_object())
                self.step_combox_list[i].addItems(list(self.driver_false.back_method_dict().keys()))
                self.step_combox_list[i].setCurrentText(test_case.step[i])
                self.page_combox_list.append(self.create_Qcombox_object())
                self.page_combox_list[i].addItems(self.sections_list)
                self.page_combox_list[i].setCurrentText(test_case.page[i])
                self.page_combox_list[i].currentIndexChanged.connect(self.page_combox_list_method)
                self.locator_name_combox_list.append(self.create_Qcombox_object())
                if self.page_combox_list[i].currentText() != "NONE":
                    self.locator_name_combox_list[i].addItems(
                        self.options_dict.get(self.page_combox_list[i].currentText()))
                    self.locator_name_combox_list[i].setCurrentText(test_case.locator_name[i])
                else:
                    self.locator_name_combox_list[i].addItem(
                        self.options_dict.get(self.page_combox_list[i].currentText()))

                self.steps_table.setCellWidget(i, 0, self.step_combox_list[i])
                self.steps_table.setCellWidget(i, 1, self.page_combox_list[i])
                self.steps_table.setCellWidget(i, 2, self.locator_name_combox_list[i])

            self.list_text_clear()

    def save_test_case_action_method(self):
        if not self.title_line_edit.text():
            QMessageBox.information(self, "提示", "请输入测试标题", QMessageBox.Ok)
            return None
        if not self.package_line_edit.text():
            QMessageBox.information(self, "提示", "请输入app包名", QMessageBox.Ok)
            return None
        if not self.activity_line_edit.text():
            QMessageBox.information(self, "提示", "请输入app入口名", QMessageBox.Ok)
            return None
        if not self.exp_line_edit.text():
            QMessageBox.information(self, "提示", "请输入预期结果", QMessageBox.Ok)
            return None
        path = QFileDialog.getSaveFileName(self, "保存测试用例",
                                               os.path.join(os.path.dirname(__file__), "test_case_object",
                                                            self.title_line_edit.text()), "*.app")
        if path[0]:
            for i in range(0, len(self.data_value_list)):
                self.data_name_text_list.append(self.data_name_list[i].text())
                self.data_value_text_list.append(self.data_value_list[i].text())
                self.data_value_two_text_list.append(self.data_value_two_list[i].text())
                self.data_value_three_text_list.append(self.data_value_three_list[i].text())
                self.data_transfer_text_list.append(self.data_transfer_list[i].text())

                self.step_combox_text_list.append(self.step_combox_list[i].currentText())
                self.page_combox_text_list.append(self.page_combox_list[i].currentText())
                self.locator_name_combox_text_list.append(self.locator_name_combox_list[i].currentText())

                test_case = TestCase(self.title_line_edit.text(), self.package_line_edit.text(),self.activity_line_edit.text(), self.step_combox_text_list,
                                     self.page_combox_text_list, self.locator_name_combox_text_list,
                                     self.data_name_text_list,
                                     self.data_value_text_list, self.data_value_two_text_list,
                                     self.data_value_three_text_list,
                                     self.data_transfer_text_list, self.exp_line_edit.text(),
                                     self.assert_method_combox.currentText())

            pickle.dump(test_case, open(path[0], "wb"))
            self.list_text_clear()

    def package_excute_method(self,driver):
        self.excute_script = BasePage(driver, self.title_line_edit.text() + ".log")
        self.excute_script.logger.info("{}_用例开始执行".format(self.title_line_edit.text()))
        self.true_dict = self.excute_script.back_method_dict()
        for i in range(0, len(self.step_combox_list)):
            if self.page_combox_list[i].currentText() == "NONE":  # 此时为浏览器没有元素定位操作和python语句操作
                para_num = len(inspect.getfullargspec(self.true_dict[self.step_combox_list[i].currentText()]).args)
                if para_num > 1:  # 有参数
                    if para_num == 2:
                        if self.data_value_list[i].text().endswith(".app"):
                            index = int(self.data_value_list[i].text().split(".")[0])
                            back_data = self.true_dict[self.step_combox_list[i].currentText()](
                                self.global_para[index])
                        else:
                            back_data = self.true_dict[self.step_combox_list[i].currentText()](
                                self.data_value_list[i].text())
                        if back_data:  # 判断是否有返回值
                            self.data_output_value_list[i].setText(back_data)
                            if self.data_transfer_list[i].text():
                                if len(eval(self.data_transfer_list[i].text())) == 3:  # 判断是否有参数传递
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                                    index = eval(self.data_transfer_list[i].text())[2]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 1:
                                    index = eval(self.data_transfer_list[i].text())[0]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 2:
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                    if para_num == 3:
                        if self.data_value_list[i].text().endswith(".app"):
                            self.data_value_list[i].setText(
                                self.global_para[int(self.data_value_list[i].text().split(".")[0])])
                        if self.data_value_two_list[i].text().endswith(".app"):
                            self.data_value_two_list[i].setText(
                                self.global_para[int(self.data_value_two_list[i].text().split(".")[0])])

                        back_data = self.true_dict[self.step_combox_list[i].currentText()](
                            self.data_value_list[i].text(),
                            self.data_value_two_list[
                                i].text())
                        if back_data:  # 判断是否有返回值
                            self.data_output_value_list[i].setText(back_data)
                            if self.data_transfer_list[i].text():
                                if len(eval(self.data_transfer_list[i].text())) == 3:  # 判断是否有参数传递
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                                    index = eval(self.data_transfer_list[i].text())[2]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 1:
                                    index = eval(self.data_transfer_list[i].text())[0]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 2:
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                    if para_num == 4:
                        if self.data_value_list[i].text().endswith(".app"):
                            self.data_value_list[i].setText(
                                self.global_para[int(self.data_value_list[i].text().split(".")[0])])
                        if self.data_value_two_list[i].text().endswith(".app"):
                            self.data_value_two_list[i].setText(
                                self.global_para[int(self.data_value_two_list[i].text().split(".")[0])])
                        if self.data_value_three_list[i].text().endswith(".app"):
                            self.data_value_three_list[i].setText(
                                self.global_para[int(self.data_value_three_list[i].text().split(".")[0])])
                        back_data = self.true_dict[self.step_combox_list[i].currentText()](
                            self.data_value_list[i].text(),
                            self.data_value_two_list[
                                i].text(),
                            self.data_value_three_list[
                                i].text())
                        if back_data:  # 判断是否有返回值
                            self.data_output_value_list[i].setText(back_data)
                            if self.data_transfer_list[i].text():
                                if len(eval(self.data_transfer_list[i].text())) == 3:  # 判断是否有参数传递
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                                    index = eval(self.data_transfer_list[i].text())[2]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 1:
                                    index = eval(self.data_transfer_list[i].text())[0]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 2:
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                else:  # 无参数
                    back_data = self.true_dict[self.step_combox_list[i].currentText()]()
                    if back_data:  # 判断是否有返回值
                        self.data_output_value_list[i].setText(back_data)
                        if self.data_transfer_list[i].text():
                            if len(eval(self.data_transfer_list[i].text())) == 3:  # 判断是否有参数传递
                                num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                    self.data_value_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                    self.data_value_two_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                    self.data_value_three_list[num].setText(back_data)
                                index = eval(self.data_transfer_list[i].text())[2]
                                self.global_para[index] = back_data
                            if len(eval(self.data_transfer_list[i].text())) == 1:
                                index = eval(self.data_transfer_list[i].text())[0]
                                self.global_para[index] = back_data
                            if len(eval(self.data_transfer_list[i].text())) == 2:
                                num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                    self.data_value_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                    self.data_value_two_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                    self.data_value_three_list[num].setText(back_data)
            else:  # 为浏览器存在元素定位操作
                para_num = len(inspect.getfullargspec(self.true_dict[self.step_combox_list[i].currentText()]).args)
                locator_name = self.cf.back_locator_tuple(self.page_combox_list[i].currentText(),
                                                          self.locator_name_combox_list[i].currentText())
                if para_num > 2:  # 有参数
                    if para_num == 3:
                        if self.data_value_list[i].text() == "USERNAME":
                            file_path = os.path.join(os.path.dirname(__file__), "LOGIN.ini")
                            with open(file_path, "r") as f:
                                read_text = f.read()
                                read_list = read_text.split(";")
                            self.true_dict[self.step_combox_list[i].currentText()](locator_name, read_list[0])
                            continue
                        if self.data_value_list[i].text() == "PASSWORD":
                            file_path = os.path.join(os.path.dirname(__file__), "LOGIN.ini")
                            with open(file_path, "r") as f:
                                read_text = f.read()
                                read_list = read_text.split(";")
                            self.true_dict[self.step_combox_list[i].currentText()](locator_name, read_list[1])
                            continue
                        if self.data_value_list[i].text().endswith(".app"):
                            self.data_value_list[i].setText(
                                self.global_para[int(self.data_value_list[i].text().split(".")[0])])
                        back_data = self.true_dict[self.step_combox_list[i].currentText()](locator_name,
                                                                                           self.data_value_list[
                                                                                               i].text())
                        if back_data:  # 判断是否有返回值
                            self.data_output_value_list[i].setText(back_data)
                            if self.data_transfer_list[i].text():
                                if len(eval(self.data_transfer_list[i].text())) == 3:  # 判断是否有参数传递
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                                    index = eval(self.data_transfer_list[i].text())[2]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 1:
                                    index = eval(self.data_transfer_list[i].text())[0]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 2:
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                    if para_num == 4:
                        if self.data_value_list[i].text().endswith(".app"):
                            self.data_value_list[i].setText(
                                self.global_para[int(self.data_value_list[i].text().split(".")[0])])
                        if self.data_value_two_list[i].text().endswith(".app"):
                            self.data_value_two_list[i].setText(
                                self.global_para[int(self.data_value_two_list[i].text().split(".")[0])])
                        back_data = self.true_dict[self.step_combox_list[i].currentText()](locator_name,
                                                                                           self.data_value_list[
                                                                                               i].text(),
                                                                                           self.data_value_two_list[
                                                                                               i].text())
                        if back_data:  # 判断是否有返回值
                            self.data_output_value_list[i].setText(back_data)
                            if self.data_transfer_list[i].text():
                                if len(eval(self.data_transfer_list[i].text())) == 3:  # 判断是否有参数传递
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                                    index = eval(self.data_transfer_list[i].text())[2]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 1:
                                    index = eval(self.data_transfer_list[i].text())[0]
                                    self.global_para[index] = back_data
                                if len(eval(self.data_transfer_list[i].text())) == 2:
                                    num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                        self.data_value_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                        self.data_value_two_list[num].setText(back_data)
                                    if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                        self.data_value_three_list[num].setText(back_data)
                    if para_num == 5:
                        if self.data_value_list[i].text().endswith(".app"):
                            self.data_value_list[i].setText(
                                self.global_para[int(self.data_value_list[i].text().split(".")[0])])
                        if self.data_value_two_list[i].text().endswith(".app"):
                            self.data_value_two_list[i].setText(
                                self.global_para[int(self.data_value_two_list[i].text().split(".")[0])])
                        if self.data_value_three_list[i].text().endswith(".app"):
                            self.data_value_three_list[i].setText(
                                self.global_para[int(self.data_value_three_list[i].text().split(".")[0])])
                        back_data = self.true_dict[self.step_combox_list[i].currentText()](locator_name,
                                                                                           self.data_value_list[
                                                                                               i].text(),
                                                                                           self.data_value_two_list[
                                                                                               i].text(),
                                                                                           self.data_value_three_list[
                                                                                               i].text())
                        if back_data:  # 判断是否有返回值
                            self.data_output_value_list[i].setText(back_data)
                            if len(eval(self.data_transfer_list[i].text())) == 3:  # 判断是否有参数传递
                                num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                    self.data_value_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                    self.data_value_two_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                    self.data_value_three_list[num].setText(back_data)
                                index = eval(self.data_transfer_list[i].text())[2]
                                self.global_para[index] = back_data
                            if len(eval(self.data_transfer_list[i].text())) == 1:
                                index = eval(self.data_transfer_list[i].text())[0]
                                self.global_para[index] = back_data
                            if len(eval(self.data_transfer_list[i].text())) == 2:
                                num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                    self.data_value_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                    self.data_value_two_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                    self.data_value_three_list[num].setText(back_data)
                else:  # 无参数
                    back_data = self.true_dict[self.step_combox_list[i].currentText()](locator_name)
                    if back_data:  # 判断是否有返回值
                        self.data_output_value_list[i].setText(back_data)
                        if self.data_transfer_list[i].text():
                            if len(eval(self.data_transfer_list[i].text())) == 3:  # 判断是否有参数传递
                                num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                    self.data_value_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                    self.data_value_two_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                    self.data_value_three_list[num].setText(back_data)
                                index = eval(self.data_transfer_list[i].text())[2]
                                self.global_para[index] = back_data
                            if len(eval(self.data_transfer_list[i].text())) == 1:
                                index = eval(self.data_transfer_list[i].text())[0]
                                self.global_para[index] = back_data
                            if len(eval(self.data_transfer_list[i].text())) == 2:
                                num = int(eval(self.data_transfer_list[i].text())[0]) - 1
                                if int(eval(self.data_transfer_list[i].text())[1]) == 1:
                                    self.data_value_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 2:
                                    self.data_value_two_list[num].setText(back_data)
                                if int(eval(self.data_transfer_list[i].text())[1]) == 3:
                                    self.data_value_three_list[num].setText(back_data)
        if self.exp_line_edit.text().endswith(".app"):
            self.exp_line_edit.setText(self.global_para[int(self.exp_line_edit.text().split(".")[0])])
        png_name = os.path.join(os.path.dirname(__file__), "test_screenshot_png", self.title_line_edit.text() + ".png")
        self.true_dict["test_screenshot_png"](png_name)

    def single_excute_action_method(self):
        self.result_label.setVisible(False)
        conf = ConfigParser()
        conf.read(filenames="SERVICEIP.ini")
        if not conf.get("deviceName_and_version","deviceName"):
            QMessageBox.information(self, "提示", "请输入手机设备名称", QMessageBox.Ok)
            return None
        if not conf.get("deviceName_and_version","host"):
            QMessageBox.information(self, "提示", "请输入测试主机IP", QMessageBox.Ok)
            return None
        if not conf.get("deviceName_and_version","version"):
            QMessageBox.information(self, "提示", "请输入手机版本", QMessageBox.Ok)
            return None
        if not conf.get("deviceName_and_version","port"):
            QMessageBox.information(self, "提示", "请输入测试主机端口", QMessageBox.Ok)
            return None
        if not self.title_line_edit.text():
            QMessageBox.information(self, "提示", "请输入测试标题", QMessageBox.Ok)
            return None
        if not self.package_line_edit.text():
            QMessageBox.information(self, "提示", "请输入app包名", QMessageBox.Ok)
            return None
        if not self.activity_line_edit.text():
            QMessageBox.information(self, "提示", "请输入app入口名", QMessageBox.Ok)
            return None
        if not self.exp_line_edit.text():
            QMessageBox.information(self, "提示", "请输入预期结果", QMessageBox.Ok)
            return None
        self.driver_test_para = BasePage(driver="driver", title="test.log").back_method_dict()
        for i in range(0, len(self.step_combox_list)):
            if self.data_transfer_list[i].text():  # 判断数据传递中格式以及返回给第几个参数的检查
                try:
                    if not isinstance(eval(self.data_transfer_list[i].text()), list):
                        QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递格式有问题，数据必须是列表！！！", QMessageBox.Ok)
                        return None
                    if len(eval(self.data_transfer_list[i].text())) == 2:
                        if eval(self.data_transfer_list[i].text())[1] not in [1, 2, 3]:
                            QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第二个值必须为数字1,2,3中的一个！！！",
                                                    QMessageBox.Ok)
                            return None
                        if not isinstance(eval(self.data_transfer_list[i].text())[0], int):
                            QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第一值必须是int类型！！！",
                                                    QMessageBox.Ok)
                            return None
                        else:
                            if eval(self.data_transfer_list[i].text())[0] < 1:
                                QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第一值必须是大于等于1的数！！！",
                                                        QMessageBox.Ok)
                                return None
                    if len(eval(self.data_transfer_list[i].text())) == 1:
                        if not isinstance(eval(self.data_transfer_list[i].text())[0], int):
                            QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第一个值必须为int类型！！！",
                                                    QMessageBox.Ok)
                            return None

                        if eval(self.data_transfer_list[i].text())[0] < 0 or eval(self.data_transfer_list[i].text())[
                            0] > 29:
                            QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第一个值必须为>=0并且<=29！！！",
                                                    QMessageBox.Ok)
                            return None
                    if len(eval(self.data_transfer_list[i].text())) == 3:
                        if not isinstance(eval(self.data_transfer_list[i].text())[2], int):
                            QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第一个值必须为int类型！！！",
                                                    QMessageBox.Ok)
                            return None

                        if eval(self.data_transfer_list[i].text())[2] < 0 or eval(self.data_transfer_list[i].text())[
                            2] > 29:
                            QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第一个值必须为>=0并且<=29！！！",
                                                    QMessageBox.Ok)
                            return None
                        if eval(self.data_transfer_list[i].text())[1] not in [1, 2, 3]:
                            QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第二个值必须为数字1,2,3中的一个！！！",
                                                    QMessageBox.Ok)
                            return None
                        if not isinstance(eval(self.data_transfer_list[i].text())[0], int):
                            QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第一值必须是int类型！！！",
                                                    QMessageBox.Ok)
                            return None

                        if eval(self.data_transfer_list[i].text())[0] < 1:
                            QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递列表中的第一值必须是大于等于1的数！！！",
                                                    QMessageBox.Ok)
                            return None
                except Exception:
                    QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的数据传递格式有问题，数据必须是列表或元组类型！！！",
                                            QMessageBox.Ok)
                    return None
                if len(eval(self.data_transfer_list[i].text())) == 2 or len(
                        eval(self.data_transfer_list[i].text())) == 3:  # 进入该条件，数据传递格式就是正确的，并给下一步设置值
                    step = int(eval(self.data_transfer_list[i].text())[0]) - 1
                    para = int(eval(self.data_transfer_list[i].text())[1])
                    if para == 1:
                        self.data_value_list[step].setText("NULL")
                    if para == 2:
                        self.data_value_two_list[step].setText("NULL")
                    if para == 3:
                        self.data_value_three_list[step].setText("NULL")

            if self.page_combox_list[i].currentText() == "NONE":  # 判断是否存在元素定位参数
                sum = 1
                if self.data_value_list[i].text():
                    sum = sum + 1
                if self.data_value_two_list[i].text():
                    sum = sum + 1
                if self.data_value_three_list[i].text():
                    sum = sum + 1
                para_num = len(
                    inspect.getfullargspec(self.driver_test_para[self.step_combox_list[i].currentText()]).args)
                if sum != para_num:
                    del sum
                    del para_num
                    QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的参数错误,请参照文档修改！！！", QMessageBox.Ok)
                    return None
                del sum
                del para_num
            else:
                sum = 2
                if self.data_value_list[i].text():
                    sum = sum + 1
                if self.data_value_two_list[i].text():
                    sum = sum + 1
                if self.data_value_three_list[i].text():
                    sum = sum + 1
                para_num = len(
                    inspect.getfullargspec(self.driver_test_para[self.step_combox_list[i].currentText()]).args)
                if sum != para_num:
                    del sum
                    del para_num
                    QMessageBox.information(self, "提示", "第" + str(i + 1) + "步的参数错误,请参照文档修改！！！", QMessageBox.Ok)
                    return None
                del sum
                del para_num
        else:
            del self.driver_test_para
        try:
            if self.chrome_radio.isChecked():
                desired_caps = {}
                desired_caps["appPackage"] = self.package_line_edit.text()
                desired_caps["appActivity"] = self.activity_line_edit.text()
                desired_caps["platformName"] = "Android"
                desired_caps["automationName"] = "uiautomator2"
                desired_caps["platformVersion"] = conf.get("deviceName_and_version","version")
                desired_caps["deviceName"] =  conf.get("deviceName_and_version","devicename")
                desired_caps["noReset"] = "True"
                desired_caps["chromedriverExecutable"]: os.path.join(os.path.dirname(__file__), "chromedriver.exe")
                desired_caps["unicodeKeyboard"] = "True"
                desired_caps['resetKeyboard'] ="True"
                host = conf.get("deviceName_and_version","host")
                port =  conf.get("deviceName_and_version","port")
                if self.ie_radio.isChecked():
                    desired_caps["chromeOptions"] = {'androidProcess': 'com.tencent.mm:tools'}
                self.driver_true = webdriver.Remote("http://"+host+":"+port+"/wd/hub", desired_caps)
            if self.firefox_radio.isChecked():
                desired_caps = {}
                desired_caps["appPackage"] = self.package_line_edit.text()
                desired_caps["appActivity"] = self.activity_line_edit.text()
                desired_caps["platformName"] = "Android"
                desired_caps["automationName"] = "uiautomator2"
                desired_caps["platformVersion"] = conf.get("deviceName_and_version", "version")
                desired_caps["deviceName"] = conf.get("deviceName_and_version", "devicename")
                desired_caps["noReset"] = "True"
                desired_caps["chromedriverExecutable"]: os.path.join(os.path.dirname(__file__), "chromedriver.exe")
                desired_caps["unicodeKeyboard"] = "True"
                desired_caps['resetKeyboard'] = "True"
                host = conf.get("deviceName_and_version", "host")
                port = conf.get("deviceName_and_version", "port")
                if self.ie_radio.isChecked():
                    desired_caps["chromeOptions"] = {'androidProcess': 'com.tencent.mm:tools'}
                self.driver_true = webdriver.Remote("http://" + host + ":" + port + "/wd/hub", desired_caps)
        except Exception as e:
            QMessageBox.information(self, "提示", "初始化失败，请检查Appium配置", QMessageBox.Ok)
            return None

        try:
            self.package_excute_method(self.driver_true)
            self.act_line_edit.setText(self.data_output_value_list[-1].text())
            if self.assert_method_combox.currentText() == "相等":
                if self.act_line_edit.text() == self.exp_line_edit.text():
                    self.result_label.setText(self.result_success)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_success + "\n")
                else:
                    self.result_label.setText(self.result_fail)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_fail + "\n")
            if self.assert_method_combox.currentText() == "不相等":
                if self.act_line_edit.text() != self.exp_line_edit.text():
                    self.result_label.setText(self.result_success)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_success + "\n")
                else:
                    self.result_label.setText(self.result_fail)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_fail + "\n")
            self.statu.showMessage("用例执行完毕")
            if self.default_teardown_value == "每个用例执行完关闭app":
                self.driver_true.quit()
        except:
            png_name = os.path.join(os.path.dirname(__file__), "test_screenshot_png",
                                    self.title_line_edit.text() + ".png")
            self.true_dict["test_screenshot_png"](png_name)
            self.driver_true.quit()
            self.result_label.setText(self.result_error)
            self.statu.showMessage("用例执行完毕")
            self.excute_script.logger.error("用例执行异常，请检查脚本\n")

    def single_excute_action_method_three(self):  # 用例失败后重跑成功删除该用例
        conf = ConfigParser()
        conf.read(filenames="SERVICEIP.ini")
        try:
            if self.chrome_radio.isChecked():
                desired_caps = {}
                desired_caps["appPackage"] = self.package_line_edit.text()
                desired_caps["appActivity"] = self.activity_line_edit.text()
                desired_caps["platformName"] = "Android"
                desired_caps["automationName"] = "uiautomator2"
                desired_caps["platformVersion"] = conf.get("deviceName_and_version", "version")
                desired_caps["deviceName"] = conf.get("deviceName_and_version", "devicename")
                desired_caps["noReset"] = "True"
                desired_caps["chromedriverExecutable"]: os.path.join(os.path.dirname(__file__), "chromedriver.exe")
                desired_caps["unicodeKeyboard"] = "True"
                desired_caps['resetKeyboard'] = "True"
                host = conf.get("deviceName_and_version", "host")
                port = conf.get("deviceName_and_version", "port")
                if self.ie_radio.isChecked():
                    desired_caps["chromeOptions"] = {'androidProcess': 'com.tencent.mm:tools'}
                self.driver_true = webdriver.Remote("http://" + host + ":" + port + "/wd/hub", desired_caps)
            if self.firefox_radio.isChecked():
                desired_caps = {}
                desired_caps["appPackage"] = self.package_line_edit.text()
                desired_caps["appActivity"] = self.activity_line_edit.text()
                desired_caps["platformName"] = "Android"
                desired_caps["automationName"] = "uiautomator2"
                desired_caps["platformVersion"] = conf.get("deviceName_and_version", "version")
                desired_caps["deviceName"] = conf.get("deviceName_and_version", "devicename")
                desired_caps["noReset"] = "True"
                desired_caps["chromedriverExecutable"]: os.path.join(os.path.dirname(__file__), "chromedriver.exe")
                desired_caps["unicodeKeyboard"] = "True"
                desired_caps['resetKeyboard'] = "True"
                host = conf.get("deviceName_and_version", "host")
                port = conf.get("deviceName_and_version", "port")
                if self.ie_radio.isChecked():
                    desired_caps["chromeOptions"] = {'androidProcess': 'com.tencent.mm:tools'}
                self.driver_true = webdriver.Remote("http://" + host + ":" + port + "/wd/hub", desired_caps)
        except Exception as e:
            QMessageBox.information(self, "提示", "初始化失败，请检查Appium配置", QMessageBox.Ok)
            return None
        start_time = time.time()
        try:
            self.package_excute_method(self.driver_true)
            self.act_line_edit.setText(self.data_output_value_list[-1].text())
            if self.assert_method_combox.currentText() == "相等":
                if self.act_line_edit.text() == self.exp_line_edit.text():
                    self.result_label.setText(self.result_success)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_success + "\n")
                    self.result_success_num += 1
                    os.remove(excute_path)
                else:
                    self.result_label.setText(self.result_fail)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_fail + "\n")
                    self.result_fail_num += 1
                    # shutil.copy(excute_path, os.path.join(os.path.dirname(__file__), "error_and_fail_test_case"))
            if self.assert_method_combox.currentText() == "不相等":
                if self.act_line_edit.text() != self.exp_line_edit.text():
                    self.result_label.setText(self.result_success)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_success + "\n")
                    self.result_success_num += 1
                    os.remove(excute_path)
                else:
                    self.result_label.setText(self.result_fail)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_fail + "\n")
                    self.result_fail_num += 1
                    # shutil.copy(excute_path, os.path.join(os.path.dirname(__file__), "error_and_fail_test_case"))
            self.statu.showMessage("用例执行完毕")
            end_time = time.time()
            if self.default_teardown_value == "每个用例执行完关闭app":
                self.driver_true.quit()
            with open(os.path.join(os.path.dirname(__file__), "test_case_log", self.title_line_edit.text() + ".log"),
                      "r", encoding="utf-8") as f:
                log_text = f.read()
            if os.path.exists(os.path.join(os.path.dirname(__file__), "test_case_report")):
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "a+") as f:
                    if self.result_label.text() == self.result_success:
                        f.write(TestHtmlReport.HTMLROWPASS.format(self.basename, self.title_line_edit.text(), log_text,
                                                                  end_time - start_time, self.result_label.text()))
                    if self.result_label.text() == self.result_fail:
                        f.write(
                            TestHtmlReport.HTMLROWUNPASS.format(self.basename, self.title_line_edit.text(), log_text,
                                                                end_time - start_time, self.result_label.text()))
            else:
                os.mkdir(os.path.join(os.path.dirname(__file__), "test_case_report"))
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "a+") as f:
                    if self.result_label.text() == self.result_success:
                        f.write(TestHtmlReport.HTMLROWPASS.format(self.basename, self.title_line_edit.text(), log_text,
                                                                  end_time - start_time, self.result_label.text()))
                    if self.result_label.text() == self.result_fail:
                        f.write(
                            TestHtmlReport.HTMLROWUNPASS.format(self.basename, self.title_line_edit.text(), log_text,
                                                                end_time - start_time, self.result_label.text()))

        except:
            # shutil.copy(excute_path, os.path.join(os.path.dirname(__file__), "error_and_fail_test_case"))
            end_time = time.time()
            png_name = os.path.join(os.path.dirname(__file__), "test_screenshot_png",
                                    self.title_line_edit.text() + ".png")
            self.true_dict["test_screenshot_png"](png_name)
            self.driver_true.quit()
            self.result_label.setText(self.result_error)
            self.statu.showMessage("用例执行完毕")
            self.excute_script.logger.error("用例执行异常，请检查脚本\n")
            self.result_error_num += 1
            with open(os.path.join(os.path.dirname(__file__), "test_case_log", self.title_line_edit.text() + ".log"),
                      "r", encoding="utf-8") as f:
                log_text = f.read()
            if os.path.exists(os.path.join(os.path.dirname(__file__), "test_case_report")):
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "a+") as f:
                    f.write(TestHtmlReport.HTMLROWERROR.format(self.basename, self.title_line_edit.text(), log_text,
                                                               end_time - start_time, self.result_label.text()))
            else:
                os.mkdir(os.path.join(os.path.dirname(__file__), "test_case_report"))
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "a+") as f:
                    f.write(TestHtmlReport.HTMLROWERROR.format(self.basename, self.title_line_edit.text(), log_text,
                                                               end_time - start_time, self.result_label.text()))

    def single_excute_action_method_two(self):  # 用例失败copy一份出来
        conf = ConfigParser()
        conf.read(filenames="SERVICEIP.ini")
        try:
            if self.chrome_radio.isChecked():
                desired_caps = {}
                desired_caps["appPackage"] = self.package_line_edit.text()
                desired_caps["appActivity"] = self.activity_line_edit.text()
                desired_caps["platformName"] = "Android"
                desired_caps["automationName"] = "uiautomator2"
                desired_caps["platformVersion"] = conf.get("deviceName_and_version", "version")
                desired_caps["deviceName"] = conf.get("deviceName_and_version", "devicename")
                desired_caps["noReset"] = "True"
                desired_caps["chromedriverExecutable"]: os.path.join(os.path.dirname(__file__), "chromedriver.exe")
                desired_caps["unicodeKeyboard"] = "True"
                desired_caps['resetKeyboard'] = "True"
                host = conf.get("deviceName_and_version", "host")
                port = conf.get("deviceName_and_version", "port")
                if self.ie_radio.isChecked():
                    desired_caps["chromeOptions"] = {'androidProcess': 'com.tencent.mm:tools'}
                self.driver_true = webdriver.Remote("http://" + host + ":" + port + "/wd/hub", desired_caps)
            if self.firefox_radio.isChecked():
                desired_caps = {}
                desired_caps["appPackage"] = self.package_line_edit.text()
                desired_caps["appActivity"] = self.activity_line_edit.text()
                desired_caps["platformName"] = "Android"
                desired_caps["automationName"] = "uiautomator2"
                desired_caps["platformVersion"] = conf.get("deviceName_and_version", "version")
                desired_caps["deviceName"] = conf.get("deviceName_and_version", "devicename")
                desired_caps["noReset"] = "True"
                desired_caps["chromedriverExecutable"]: os.path.join(os.path.dirname(__file__), "chromedriver.exe")
                desired_caps["unicodeKeyboard"] = "True"
                desired_caps['resetKeyboard'] = "True"
                host = conf.get("deviceName_and_version", "host")
                port = conf.get("deviceName_and_version", "port")
                if self.ie_radio.isChecked():
                    desired_caps["chromeOptions"] = {'androidProcess': 'com.tencent.mm:tools'}
                self.driver_true = webdriver.Remote("http://" + host + ":" + port + "/wd/hub", desired_caps)
        except Exception as e:
            QMessageBox.information(self, "提示", "初始化失败，请检查Appium配置", QMessageBox.Ok)
            return None
        start_time = time.time()
        try:
            self.package_excute_method(self.driver_true)
            self.act_line_edit.setText(self.data_output_value_list[-1].text())
            if self.assert_method_combox.currentText() == "相等":
                if self.act_line_edit.text() == self.exp_line_edit.text():
                    self.result_label.setText(self.result_success)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_success + "\n")
                    self.result_success_num += 1
                    # os.remove(excute_path)
                else:
                    self.result_label.setText(self.result_fail)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_fail + "\n")
                    self.result_fail_num += 1
                    shutil.copy(excute_path, os.path.join(os.path.dirname(__file__), "error_and_fail_test_case"))
            if self.assert_method_combox.currentText() == "不相等":
                if self.act_line_edit.text() != self.exp_line_edit.text():
                    self.result_label.setText(self.result_success)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_success + "\n")
                    self.result_success_num += 1
                    # os.remove(excute_path)
                else:
                    self.result_label.setText(self.result_fail)
                    self.excute_script.logger.info("用例的执行情况是:" + self.result_fail + "\n")
                    self.result_fail_num += 1
                    shutil.copy(excute_path, os.path.join(os.path.dirname(__file__), "error_and_fail_test_case"))
            self.statu.showMessage("用例执行完毕")
            end_time = time.time()
            if self.default_teardown_value == "每个用例执行完关闭app":
                self.driver_true.quit()
            with open(os.path.join(os.path.dirname(__file__), "test_case_log", self.title_line_edit.text() + ".log"),
                      "r", encoding="utf-8") as f:
                log_text = f.read()
            if os.path.exists(os.path.join(os.path.dirname(__file__), "test_case_report")):
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "a+") as f:
                    if self.result_label.text() == self.result_success:
                        f.write(TestHtmlReport.HTMLROWPASS.format(self.basename, self.title_line_edit.text(), log_text,
                                                                  end_time - start_time, self.result_label.text()))
                    if self.result_label.text() == self.result_fail:
                        f.write(
                            TestHtmlReport.HTMLROWUNPASS.format(self.basename, self.title_line_edit.text(), log_text,
                                                                end_time - start_time, self.result_label.text()))
            else:
                os.mkdir(os.path.join(os.path.dirname(__file__), "test_case_report"))
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "a+") as f:
                    if self.result_label.text() == self.result_success:
                        f.write(TestHtmlReport.HTMLROWPASS.format(self.basename, self.title_line_edit.text(), log_text,
                                                                  end_time - start_time, self.result_label.text()))
                    if self.result_label.text() == self.result_fail:
                        f.write(
                            TestHtmlReport.HTMLROWUNPASS.format(self.basename, self.title_line_edit.text(), log_text,
                                                                end_time - start_time, self.result_label.text()))

        except:
            shutil.copy(excute_path, os.path.join(os.path.dirname(__file__), "error_and_fail_test_case"))
            end_time = time.time()
            png_name = os.path.join(os.path.dirname(__file__), "test_screenshot_png",
                                    self.title_line_edit.text() + ".png")
            self.true_dict["test_screenshot_png"](png_name)
            self.driver_true.quit()
            self.result_label.setText(self.result_error)
            self.statu.showMessage("用例执行完毕")
            self.excute_script.logger.error("用例执行异常，请检查脚本\n")
            self.result_error_num += 1
            with open(os.path.join(os.path.dirname(__file__), "test_case_log", self.title_line_edit.text() + ".log"),
                      "r", encoding="utf-8") as f:
                log_text = f.read()
            if os.path.exists(os.path.join(os.path.dirname(__file__), "test_case_report")):
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "a+") as f:
                    f.write(TestHtmlReport.HTMLROWERROR.format(self.basename, self.title_line_edit.text(), log_text,
                                                               end_time - start_time, self.result_label.text()))
            else:
                os.mkdir(os.path.join(os.path.dirname(__file__), "test_case_report"))
                with open(
                        os.path.join(os.path.dirname(__file__), "test_case_report", self.excute_time + "_report.html"),
                        "a+") as f:
                    f.write(TestHtmlReport.HTMLROWERROR.format(self.basename, self.title_line_edit.text(), log_text,
                                                               end_time - start_time, self.result_label.text()))

    def sub_step_btn_method(self):
        if self.steps_table_row >= 2:
            self.steps_table_row = self.steps_table_row - 1
            self.table_row = self.table_row - 1

            self.step_combox_list.pop(-1)
            self.page_combox_list.pop(-1)
            self.locator_name_combox_list.pop(-1)

            self.data_name_list.pop(-1)
            self.data_value_list.pop(-1)
            self.data_value_two_list.pop(-1)
            self.data_value_three_list.pop(-1)
            self.data_output_value_list.pop(-1)
            self.data_transfer_list.pop(-1)

            self.steps_table.setRowCount(self.steps_table_row)
            self.data_table.setRowCount(self.table_row)

    def email_action_method(self):
        file_path = os.path.join(os.path.dirname(__file__), "EMAIL.ini")
        conf = ConfigParser()
        conf.read(filenames="EMAIL.ini")
        dialog = QDialog()
        dialog.setWindowTitle("邮箱配置")
        email_service_edit = QLineEdit(dialog)
        email_service_edit.setPlaceholderText("请输入邮箱服务器名称")
        email_service_edit.resize(150, 30)
        email_true_radio = QRadioButton("生效", dialog)
        email_false_radio = QRadioButton("不生效", dialog)

        def email_true_method():
            self.email_status = True

        def email_false_method():
            self.email_status = False

        email_true_radio.toggled.connect(email_true_method)
        email_false_radio.toggled.connect(email_false_method)
        if self.email_status:
            email_true_radio.setChecked(True)
            self.email_status = True
        else:
            email_false_radio.setChecked(True)
            self.email_status = False
        email_false_radio.move(170, 10)
        email_true_radio.move(245, 10)
        send_user_name_edit = QLineEdit(dialog)
        send_password_edit = QLineEdit(dialog)
        receve_name_one = QLineEdit(dialog)
        receve_name_two = QLineEdit(dialog)
        receve_name_three = QLineEdit(dialog)
        send_user_name_edit.resize(300, 30)
        send_user_name_edit.move(0, 30)
        email_service_edit.setText(conf.get("send", "email_service"))
        send_user_name_edit.setText(conf.get("send", "user_name"))
        send_user_name_edit.setPlaceholderText("请输入发送人的邮箱用户名")
        send_password_edit.resize(300, 30)
        send_password_edit.setText(conf.get("send", "password"))
        send_password_edit.move(0, 60)
        send_password_edit.setEchoMode(QLineEdit.Password)
        send_password_edit.setPlaceholderText("请输入发送人的邮箱密码")
        receve_name_one.resize(300, 30)
        receve_name_one.setText(conf.get("receve", "receve_user_one"))
        receve_name_two.resize(300, 30)
        receve_name_two.setText(conf.get("receve", "receve_user_two"))
        receve_name_three.resize(300, 30)
        receve_name_three.setText(conf.get("receve", "receve_user_three"))
        receve_name_three.setPlaceholderText("请输入接受者邮箱")
        receve_name_two.setPlaceholderText("请输入接受者邮箱")
        receve_name_one.setPlaceholderText("请输入接受者邮箱")
        receve_name_one.move(0, 90)
        receve_name_two.move(0, 120)
        receve_name_three.move(0, 150)

        def remove_btn_method():
            dialog.close()

        def submit_btn_method():
            with open(file_path, "w+") as f:
                conf = ConfigParser()
                conf.read(filenames="EMAIL.ini")
                conf.add_section("send")
                conf.set("send", "email_service", email_service_edit.text())
                conf.set("send", "user_name", send_user_name_edit.text())
                conf.set("send", "password", send_password_edit.text())
                conf.add_section("receve")
                conf.set("receve", "receve_user_one", receve_name_one.text())
                conf.set("receve", "receve_user_two", receve_name_two.text())
                conf.set("receve", "receve_user_three", receve_name_three.text())
                conf.write(f)
            dialog.close()

        submit_btn = QPushButton("确定", dialog)
        submit_btn.clicked.connect(submit_btn_method)
        remove_btn = QPushButton("取消", dialog)
        remove_btn.clicked.connect(remove_btn_method)
        submit_btn.resize(150, 30)
        submit_btn.move(0, 180)
        remove_btn.resize(150, 30)
        remove_btn.move(150, 180)
        dialog.resize(300, 210)
        dialog.exec_()

    def test_service_url_method(self):
        file_path = os.path.join(os.path.dirname(__file__), "SERVICEIP.ini")
        conf = ConfigParser()
        conf.read(filenames="SERVICEIP.ini")
        dialog = QDialog(self)
        dialog.resize(300, 90)
        dialog.setWindowTitle("设备名称和手机版本配置")
        self.device_line_edit = QLineEdit(dialog)
        self.version_line_edit = QLineEdit(dialog)
        self.host_line_edit = QLineEdit(dialog)
        self.port_line_edit = QLineEdit(dialog)
        self.device_line_edit.setPlaceholderText("请输入设备名称")
        self.version_line_edit.setPlaceholderText("请输入手机版本号")
        self.host_line_edit.setPlaceholderText("请输入主机IP")
        self.port_line_edit.setPlaceholderText("请输入主机端口")

        def remove_btn_method():
            dialog.close()

        def submit_btn_method():
            with open(file_path, "w+") as f:
                conf = ConfigParser()
                conf.read(filenames="SERVICEIP.ini")
                conf.add_section("deviceName_and_version")
                conf.set("deviceName_and_version", "deviceName", self.device_line_edit.text())
                conf.set("deviceName_and_version", "version", self.version_line_edit.text())
                conf.set("deviceName_and_version", "host", self.host_line_edit.text())
                conf.set("deviceName_and_version", "port", self.port_line_edit.text())
                conf.write(f)
            dialog.close()

        submit_btn = QPushButton("确定", dialog)
        submit_btn.clicked.connect(submit_btn_method)
        remove_btn = QPushButton("取消", dialog)
        remove_btn.clicked.connect(remove_btn_method)
        submit_btn.resize(150, 30)
        submit_btn.move(0, 60)
        remove_btn.resize(150, 30)
        remove_btn.move(150, 60)
        self.device_line_edit.resize(150, 30)
        self.version_line_edit.resize(150,30)
        self.host_line_edit.resize(150,30)
        self.port_line_edit.resize(150,30)
        self.version_line_edit.move(0,30)
        self.host_line_edit.move(150,0)
        self.port_line_edit.move(150,30)
        self.device_line_edit.setText(conf.get("deviceName_and_version","deviceName"))
        self.version_line_edit.setText(conf.get("deviceName_and_version","version"))
        self.host_line_edit.setText(conf.get("deviceName_and_version", "host"))
        self.port_line_edit.setText(conf.get("deviceName_and_version", "port"))
        dialog.exec_()

    def teardown_action_method(self):
        tear_down_list = ["每个用例执行完关闭app", "无操作"]
        items = QInputDialog.getItem(self, "后置处理选择", "后置处理选择", tear_down_list,
                                     tear_down_list.index(self.default_teardown_value), False)
        if items[1]:
            self.default_teardown_value = items[0]
        else:
            pass

    def user_information_method(self):
        dialog = QDialog()
        dialog.setWindowTitle("按钮使用说明")
        user_text = QTextEdit(dialog)
        user_text.setReadOnly(True)
        user_text.resize(QApplication.desktop().width(), QApplication.desktop().height())
        user_text.setText("""一  分辨率以及程序初始化说明

     默认支持所有分辨率

     程序打开时，会判断同目录下是否存在LocatorsObject.ini文件，该文件存放元素定位的文件，请统一在里面配置，格式请参照默认配置

     执行测试用例需要把对应的浏览器驱动放在同exe程序同目录下。


二  具体按钮说明

     全局配置中的【服务器IP及端口配置】主要是为了测试服务器域名更换时，对批量测试脚本的影响。


    *域名填写的时候后面必须在后面加上”/”符号，在测试网址编辑框中只需要输入具体的测试路径就可以了。


    全局配置中的【登录账号信息配置】主要是为了测试用例存在登录的情况，在登录账号信息配置中输入具体的账号和密码，在具体步骤调用时只需要在参数中输入”USERNAME”和”PASSWORD”即可。


    全局配置中的【数据库配置】是查询数据库的数据作为参数传入的场景（比如注册场景验证码的输入），只要在具体的域名，及账号，密码，端口，数据库名配置即可，在某一步调用的时候会自动读取配置文件的数据


    文件中的【保存测试用例】是将某一测试用例中的，测试标题，测试网址，预期结果，参数一，参数二，参数三，数据传递，以及所在界面，元素定位名称，和关键字步骤通过序列化保存以”.ZQC”结尾的文件中。


    文件中的【打开测试用例】就是将保存的数据通过反序列化”.ZQC”结尾的数据在展示在自动化测试平台中。


    工具栏中的【添加前置条件】主要是为了解决登录操作，只要打开”登录.ZQC”文件即可。


    工具栏中的【执行单个测试用例】主要是为了脚本调试，执行完毕可在查看日志中查看具体的执行情况，执行前还会检查每一步对应的关键字步骤对应的参数是否正确。


   【数据传递】 分为3种情况：
    第一种(只有全局传参)：在文本框中输入一个长度的列表，内容是1-30，如[1]在调用的地方输入1.ZQC。
    第二种（只有局部传参）：在文本框中输入两个长度的列表，内容是[第几步，第几个参数]，如[3，3]表示在第三步的第3个参数传入
    第三种（都有）：在文本框中输入三个长度的列表，最后一个表示全局传参。


    *每一步的最多包含三个参数，必须按从左到右依次传递，参数二和参数三可编辑的前提分别是参数一中存在值和参数二中存在值，否则就算有值，系统也会自动清空，


    *输出数据是在执行用例，系统会自动判断每个步骤有没有返回值，有就会把返回值设置在对应的【输出数据中】,点击输出数据，可以动态插入和删除测试步骤


    *全局变量传递,如果你希望某一步骤的返回数据作为其他用例的输入步骤，在返回步骤中的数据传递输入框中输入(n,),必须是英文状态下且n>=0andn<=29,
     在需要传递的数据中输入n.ZQC即可,


    *关于执行先后的设定：可以在测试标题前面加：  字母_测试标题，windows系统自动遍历按字母的ascall码值。


    *其他关键字封装以及业务逻辑封装可以加我微信：zqc18273569617
        """)
        dialog.showMaximized()
        dialog.exec_()

    def view_log_action_method(self):
        path = QFileDialog.getOpenFileName(self, "日志查看", os.path.join(os.path.dirname(__file__), "test_case_log",
                                                                          self.title_line_edit.text() + ".log"),
                                               "*.log")
        if path[0]:
            dialog = QDialog()
            dialog.setWindowTitle(os.path.basename(path[0]) + "的执行日志")
            self.text_log = QTextEdit(dialog)
            self.text_log.resize(QApplication.desktop().width(), QApplication.desktop().height())
            with open(path[0], "r+", encoding="utf-8") as f:
                text = f.read()
            self.text_log.setText(text)
            dialog.showMaximized()
            dialog.exec_()

    def view_result_action_method(self):
        if self.result_label.text():
            self.result_label.setVisible(True)
        else:
            QMessageBox.information(self, "提示", "请先执行测试用例", QMessageBox.Ok)


    def create_all_dir(self):
        if not os.path.exists(os.path.join(os.path.dirname(__file__), "error_and_fail_test_case")):
            os.makedirs(os.path.join(os.path.dirname(__file__), "error_and_fail_test_case"))
        photo_dir = os.path.join(os.path.dirname(__file__), "test_screenshot_png")
        if not os.path.exists(photo_dir):
            os.mkdir(photo_dir)
        if not os.path.exists(os.path.join(os.path.dirname(__file__), "test_case_log")):
            os.mkdir(os.path.join(os.path.dirname(__file__), "test_case_log"))
        if not os.path.exists(os.path.join(os.path.dirname(__file__), "test_case_object")):
            os.mkdir(os.path.join(os.path.dirname(__file__), "test_case_object"))
        if not os.path.exists(os.path.join(os.path.dirname(__file__), "test_case_report")):
            os.mkdir(os.path.join(os.path.dirname(__file__), "test_case_report"))
        file_path_email = os.path.join(os.path.dirname(__file__), "EMAIL.ini")
        if not os.path.exists(file_path_email):
            with open(file_path_email, "w+") as f:
                conf = ConfigParser()
                conf.read(filenames="EMAIL.ini")
                conf.add_section("send")
                conf.set("send", "email_service", "")
                conf.set("send", "user_name", "")
                conf.set("send", "password", "")
                conf.add_section("receve")
                conf.set("receve", "receve_user_one", "")
                conf.set("receve", "receve_user_two", "")
                conf.set("receve", "receve_user_three", "")
                conf.write(f)
        file_path_login = os.path.join(os.path.dirname(__file__), "LOGIN.ini")
        if not os.path.exists(file_path_login):
            with open(file_path_login, "w+") as f:
                pass
        file_path_email = os.path.join(os.path.dirname(__file__), "MYSQL.ini")
        if not os.path.exists(file_path_email):
            with open(file_path_email, "w+") as f:
                pass
        file_path = os.path.join(os.path.dirname(__file__), "SERVICEIP.ini")
        if not os.path.exists(file_path):
            with open(file_path, "w+") as f:
                conf = ConfigParser()
                conf.read(filenames="SERVICEIP.ini")
                conf.add_section("deviceName_and_version")
                conf.set("deviceName_and_version", "deviceName", "")
                conf.set("deviceName_and_version", "version", "")
                conf.set("deviceName_and_version", "host", "")
                conf.set("deviceName_and_version", "port", "")
                conf.write(f)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    initmain = MainInit()
    initmain.create_all_dir()
    sys.exit(app.exec_())

